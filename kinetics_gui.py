#!/usr/bin/env python3
"""
kinetics_gui.py

Tkinter GUI on top of `kinetics_pipeline.py` for batch processing of multiple
enzyme-variant Cary 60 CSV files. Workflow:

  1. User selects a folder containing one CSV per variant.
  2. Slopes are extracted per file (via kinetics_pipeline.analyse_file).
  3. A "Setup" tab shows an editable table of enzymes (one row per file)
     where the user enters [E] (uM) and epsilon (M^-1 cm^-1) per variant.
  4. A tab per variant lists the slope pairs with editable per-pair
     [S] (uM) and an "include" checkbox. Default concentrations follow
     the user's convention (5 levels x 3 reps).
  5. A "Plot" tab overlays kobs vs [S] for all enzymes.
  6. "Export" writes a per-pair detail CSV and a per-enzyme summary CSV.

Run:
    python kinetics_gui.py [folder]

Requires `kinetics_pipeline.py` to be importable from the same directory
(i.e. on PYTHONPATH or sitting next to this script).

Author: Cornel Niederhauser, 2026
"""

from __future__ import annotations

import csv
import sys
import traceback
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk,
)

# Local pipeline module
import kinetics_pipeline as kp


# ---------------------------------------------------------------------------
# Constants & defaults
# ---------------------------------------------------------------------------

DEFAULT_EPSILON_M1CM1 = 15784.925513164
DEFAULT_PATHLENGTH_CM = 1.0
DEFAULT_ENZYME_CONC_UM = 1.0          # placeholder; user must edit per variant
DEFAULT_CONCENTRATIONS_UM = [
    20.0,  20.0,  20.0,
    37.5,  37.5,  37.5,
    75.0,  75.0,  75.0,
   150.0, 150.0, 150.0,
   300.0, 300.0, 300.0,
]


# ---------------------------------------------------------------------------
# Calculation helpers
# ---------------------------------------------------------------------------

def velocity_uM_per_s(slope_AU_per_min: float, epsilon_M1cm1: float,
                      pathlength_cm: float = DEFAULT_PATHLENGTH_CM) -> float:
    """v [uM/s] = slope_net [AU/min] / (l * eps / 1e6) / 60"""
    if epsilon_M1cm1 <= 0 or pathlength_cm <= 0:
        return float("nan")
    return slope_AU_per_min / (pathlength_cm * epsilon_M1cm1 / 1.0e6) / 60.0


def kobs_per_s(v_uM_per_s: float, enzyme_conc_uM: float) -> float:
    """kobs [1/s] = v [uM/s] / [E] [uM]."""
    if enzyme_conc_uM <= 0:
        return float("nan")
    return v_uM_per_s / enzyme_conc_uM


def fit_kcat_over_km(conc_uM: list[float], kobs_s: list[float]
                     ) -> tuple[float, float, float, int]:
    """Free OLS linear regression of kobs (s^-1) on [S] (uM).
    Returns (kcat_over_km_M1s1, intercept_s1, r2, n_used)."""
    x = np.asarray(conc_uM, dtype=float)
    y = np.asarray(kobs_s, dtype=float)
    finite = np.isfinite(x) & np.isfinite(y)
    x, y = x[finite], y[finite]
    n = len(x)
    if n < 2:
        return float("nan"), float("nan"), float("nan"), n
    slope_uM, intercept = np.polyfit(x, y, 1)
    yhat = slope_uM * x + intercept
    ss_res = float(np.sum((y - yhat) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    return float(slope_uM * 1.0e6), float(intercept), float(r2), n


def _build_replicate_groups(conc_uM: list[float], kobs_s: list[float],
                             pair_order: list[int]
                             ) -> list[list[tuple[float, float]]]:
    """Map a flat list of pairs to a list of replicate groups.

    Mapping rule: sort pairs by concentration, then within each
    concentration assign the k-th occurrence (in input/pair order) to
    replicate k. This works for both interleaved triplicates
    (`20, 37.5, 75, ..., 20, 37.5, 75, ...`) and contiguous triplicates
    (`20, 20, 20, 37.5, 37.5, ...`) since both produce three
    independent measurements per concentration.

    Returns a list of replicate groups; each group is a list of
    `(conc, kobs)` tuples for the points belonging to that replicate.
    Replicates may have unequal lengths if the data is asymmetric.
    """
    from collections import defaultdict
    # Bucket by concentration, preserving input order for the k-th-occurrence
    # rule
    order = sorted(range(len(conc_uM)),
                   key=lambda i: (conc_uM[i], pair_order[i]))
    by_conc: dict[float, list[float]] = defaultdict(list)
    for i in order:
        by_conc[conc_uM[i]].append(kobs_s[i])
    max_reps = max((len(v) for v in by_conc.values()), default=0)
    reps: list[list[tuple[float, float]]] = [[] for _ in range(max_reps)]
    for c in sorted(by_conc.keys()):
        for i, k in enumerate(by_conc[c]):
            reps[i].append((c, k))
    return reps


def fit_kcat_over_km_weighted(
    conc_uM: list[float], kobs_s: list[float],
) -> tuple[float, float, float, float, int]:
    """Replicate-aware weighted linear regression of kobs on [S].

    For each unique [S] with three or more pairs, computes the
    triplicate SD and uses it as the measurement uncertainty (weight =
    1/sigma^2). Singletons and pairs at concentrations with fewer than
    three replicates get a fallback sigma equal to the residual standard
    deviation of a preliminary OLS fit — i.e., they are treated with
    the assay-wide average noise level.

    Returns (slope_M1s1, slope_SE_M1s1, intercept_s1, r2, n_used).
    """
    x = np.asarray(conc_uM, dtype=float)
    y = np.asarray(kobs_s, dtype=float)
    finite = np.isfinite(x) & np.isfinite(y)
    x, y = x[finite], y[finite]
    n = len(x)
    if n < 3:
        slope, intercept, r2, _ = fit_kcat_over_km(conc_uM, kobs_s)
        return slope, float("nan"), intercept, r2, n

    slope_uM, intercept = np.polyfit(x, y, 1)
    resid = y - (slope_uM * x + intercept)
    sigma_pool = float(np.sqrt(np.sum(resid ** 2) / max(n - 2, 1)))
    if sigma_pool <= 0 or not np.isfinite(sigma_pool):
        sigma_pool = 1.0

    sigmas = np.full(n, sigma_pool, dtype=float)
    for c in np.unique(x):
        idx = np.where(x == c)[0]
        if len(idx) >= 3:
            sd = float(np.std(y[idx], ddof=1))
            sigmas[idx] = max(sd, 0.1 * sigma_pool)

    try:
        from scipy.optimize import curve_fit
        def line(xx, m, b):
            return m * xx + b
        popt, pcov = curve_fit(
            line, x, y, sigma=sigmas, absolute_sigma=True,
            p0=(slope_uM, intercept))
        slope_w, intercept_w = popt
        slope_se_uM = float(np.sqrt(pcov[0, 0])) if pcov[0, 0] > 0 else float("nan")
    except Exception:
        slope_w, intercept_w = slope_uM, intercept
        slope_se_uM = float("nan")

    yhat = slope_w * x + intercept_w
    ss_res = float(np.sum((y - yhat) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    return (
        float(slope_w * 1.0e6),
        float(slope_se_uM * 1.0e6) if np.isfinite(slope_se_uM) else float("nan"),
        float(intercept_w),
        float(r2),
        n,
    )


def fit_kcat_over_km_per_replicate(
    conc_uM: list[float], kobs_s: list[float], pair_order: list[int],
) -> tuple[float, float, float, float, int, int]:
    """Fit k_cat/K_M separately for each technical replicate.

    Returns (mean_slope_M1s1, sd_slope_M1s1, mean_intercept,
             mean_r2, n_replicates_used, n_total_points).

    Each replicate gets one independent OLS fit of kobs vs [S]. The
    reported value is the mean across replicates and the SD is the
    sample SD (ddof=1). Replicates with fewer than 2 points are
    skipped. Returns NaN SD if fewer than 2 replicates contribute.
    """
    x = np.asarray(conc_uM, dtype=float)
    y = np.asarray(kobs_s, dtype=float)
    finite = np.isfinite(x) & np.isfinite(y)
    x_list = x[finite].tolist()
    y_list = y[finite].tolist()
    order = [pair_order[i] for i, f in enumerate(finite) if f]
    n_total = len(x_list)
    if n_total < 4:  # need at least two points in two reps
        slope, intercept, r2, _ = fit_kcat_over_km(conc_uM, kobs_s)
        return slope, float("nan"), intercept, r2, 0, n_total

    reps = _build_replicate_groups(x_list, y_list, order)

    slopes: list[float] = []
    intercepts: list[float] = []
    r2s: list[float] = []
    for rep in reps:
        if len(rep) < 2:
            continue
        xs = np.asarray([r[0] for r in rep])
        ys = np.asarray([r[1] for r in rep])
        if len(np.unique(xs)) < 2:
            continue
        sl, ic = np.polyfit(xs, ys, 1)
        yhat = sl * xs + ic
        ss_res = float(np.sum((ys - yhat) ** 2))
        ss_tot = float(np.sum((ys - ys.mean()) ** 2))
        r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
        slopes.append(sl * 1e6)
        intercepts.append(ic)
        r2s.append(r2)

    n_reps = len(slopes)
    if n_reps < 2:
        slope, intercept, r2, _ = fit_kcat_over_km(conc_uM, kobs_s)
        return slope, float("nan"), intercept, r2, n_reps, n_total

    mean = float(np.mean(slopes))
    sd = float(np.std(slopes, ddof=1))
    mean_ic = float(np.mean(intercepts))
    mean_r2 = float(np.mean(r2s))
    return mean, sd, mean_ic, mean_r2, n_reps, n_total


def fit_kcat_over_km_bootstrap(
    conc_uM: list[float], kobs_s: list[float], pair_order: list[int],
    n_iter: int = 10000, seed: int = 0,
) -> tuple[float, float, float, float, float, int, int]:
    """Replicate-clustered bootstrap of kcat/KM with 95% CI.

    Returns (point_slope_M1s1, ci_lo_M1s1, ci_hi_M1s1,
             intercept_s1, r2, n_replicates, n_total_points).

    Bootstrap procedure: treat each technical replicate as one cluster.
    At each iteration, draw `n_reps` replicates with replacement from
    the available pool, pool all their points, fit by OLS. Report the
    OLS point estimate from the full data and the [2.5%, 97.5%]
    percentile CI from the bootstrap distribution.

    Note: with only 3 replicates there are exactly 10 unique unordered
    resample combinations, so the bootstrap distribution is discrete
    and the CI is jagged. This is a fundamental limitation of small-n
    cluster bootstrap, not a bug.
    """
    x = np.asarray(conc_uM, dtype=float)
    y = np.asarray(kobs_s, dtype=float)
    finite = np.isfinite(x) & np.isfinite(y)
    x_list = x[finite].tolist()
    y_list = y[finite].tolist()
    order = [pair_order[i] for i, f in enumerate(finite) if f]
    n_total = len(x_list)
    # Need at least 4 points for replicate-clustered bootstrap
    if n_total < 4:
        slope, intercept, r2, _ = fit_kcat_over_km(conc_uM, kobs_s)
        return (slope, float("nan"), float("nan"), intercept, r2,
                0, n_total)

    reps = _build_replicate_groups(x_list, y_list, order)
    reps = [r for r in reps if len(r) >= 2]
    n_reps = len(reps)
    if n_reps < 2:
        slope, intercept, r2, _ = fit_kcat_over_km(conc_uM, kobs_s)
        return (slope, float("nan"), float("nan"), intercept, r2,
                n_reps, n_total)

    # Point estimate on the full pooled data
    xs_all = np.concatenate([np.asarray([p[0] for p in r]) for r in reps])
    ys_all = np.concatenate([np.asarray([p[1] for p in r]) for r in reps])
    sl_pt, ic_pt = np.polyfit(xs_all, ys_all, 1)
    yhat = sl_pt * xs_all + ic_pt
    ss_res = float(np.sum((ys_all - yhat) ** 2))
    ss_tot = float(np.sum((ys_all - ys_all.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    point_slope_M1s1 = float(sl_pt * 1e6)

    # Bootstrap
    rng = np.random.default_rng(seed)
    slopes = np.empty(n_iter, dtype=float)
    n_kept = 0
    for _ in range(n_iter):
        chosen = rng.integers(0, n_reps, size=n_reps)
        xs_list = []
        ys_list = []
        for r_idx in chosen:
            for c, k in reps[r_idx]:
                xs_list.append(c)
                ys_list.append(k)
        xs_b = np.asarray(xs_list)
        ys_b = np.asarray(ys_list)
        if len(np.unique(xs_b)) < 2:
            continue
        sl_b, _ = np.polyfit(xs_b, ys_b, 1)
        slopes[n_kept] = sl_b * 1e6
        n_kept += 1
    if n_kept < 10:
        return (point_slope_M1s1, float("nan"), float("nan"),
                float(ic_pt), r2, n_reps, n_total)
    slopes = slopes[:n_kept]
    ci_lo = float(np.percentile(slopes, 2.5))
    ci_hi = float(np.percentile(slopes, 97.5))
    return (point_slope_M1s1, ci_lo, ci_hi, float(ic_pt), r2,
            n_reps, n_total)


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class PipelineParams:
    """All knobs from kinetics_pipeline that the GUI exposes."""
    min_seg_sec: float = 20.0
    head_trim_sec: float = 0.0
    tail_trim_sec: float = 1.0
    r2_thresh: float = 0.995
    slope_tol: float = 0.10
    min_window_sec: float = 8.0
    extend_to_minimum: bool = True
    extend_lookback_sec: float = 3.0
    extend_smooth_thresh: float = 0.02
    # Method for selecting the linear region of an enzyme segment.
    # 'curvature' (default): noise-robust quadratic-curvature test.
    # 'r2_window':           older R^2-based window growth.
    fit_method: str = "curvature"
    # Curvature-method specific:
    max_slope_dev: float = 0.05
    curvature_sig_sigma: float = 4.0
    # Method for combining per-pair k_obs into one k_cat/K_M per variant.
    # 'weighted':      replicate-aware weighted regression (default).
    # 'per_replicate': fit each technical replicate separately, then
    #                  report mean ± SD across replicates.
    # 'bootstrap':     replicate-clustered bootstrap with 95% CI.
    kcat_method: str = "weighted"
    bootstrap_iter: int = 10000


@dataclass
class PairRow:
    pair: int
    t_enzyme_start: float
    slope_blank: float
    slope_enzyme: float
    slope_net: float
    r2_enzyme: float
    meets_r2: bool
    window_enzyme_s: float
    # User-editable
    concentration_uM: float = float("nan")
    include: bool = True
    # Manual override of the fit window. When `manually_adjusted` is True,
    # `t_enzyme_start` and `window_enzyme_s` reflect the user-set window;
    # the slope/R^2 fields are recomputed by `_refit_pair_window` over
    # exactly those bounds (no auto extend, no R^2/slope-tolerance
    # search). The original auto-fit values are stashed in `_auto_*` so
    # we can revert.
    manually_adjusted: bool = False
    _auto_t_enzyme_start: float = float("nan")
    _auto_window_enzyme_s: float = float("nan")
    _auto_slope_enzyme: float = float("nan")
    _auto_r2_enzyme: float = float("nan")
    _auto_slope_net: float = float("nan")
    _auto_meets_r2: bool = True
    # Reference to the cached enzyme-segment slice (filled at extraction
    # time) so the plot and the manual refitter can re-fit without going
    # back through segmentation.
    _enzyme_seg: tuple[int, int] | None = None
    _blank_seg: tuple[int, int] | None = None
    # The blank fit may exclude a post-spike transient at the head of
    # the blank segment. `_blank_fit_seg` is the actual (start, end)
    # index range used for the fit; `_blank_intercept` is the fit's
    # y-intercept in the absolute (unshifted) absorbance frame.
    _blank_fit_seg: tuple[int, int] | None = None
    _blank_intercept: float = float("nan")


@dataclass
class SegmentSpec:
    """One contiguous time region tagged blank or enzyme. The user-facing
    primitive in step 1; pairs are derived from a list of these."""
    s_idx: int           # absolute index into the raw time array
    e_idx: int           # absolute index, exclusive
    label: str           # 'blank' or 'enzyme'

    def to_dict(self) -> dict:
        return {"s_idx": int(self.s_idx),
                "e_idx": int(self.e_idx),
                "label": self.label}

    @classmethod
    def from_dict(cls, d: dict) -> "SegmentSpec":
        return cls(s_idx=int(d["s_idx"]),
                   e_idx=int(d["e_idx"]),
                   label=str(d["label"]))


@dataclass
class FileSession:
    path: Path
    name: str
    pairs: list[PairRow] = field(default_factory=list)
    enzyme_conc_uM: float = DEFAULT_ENZYME_CONC_UM
    epsilon_M1cm1: float = DEFAULT_EPSILON_M1CM1
    error: Optional[str] = None
    # Computed (filled by recompute):
    kcat_over_km: float = float("nan")
    kcat_over_km_se: float = float("nan")
    # For bootstrap: half-widths to the 2.5 and 97.5 percentile bounds
    # (i.e., kcat ± these gives the CI). NaN unless kcat_method is
    # 'bootstrap' or the method explicitly fills them.
    kcat_ci_lo: float = float("nan")
    kcat_ci_hi: float = float("nan")
    # Which method actually produced the value above. Used by the
    # display strip and the Excel exporter to label uncertainties
    # correctly even after a method change.
    kcat_method_used: str = "weighted"
    fit_intercept: float = float("nan")
    fit_r2: float = float("nan")
    n_used: int = 0
    # Step 1 state: list of (s, e, label) regions, plus a confirmation
    # flag. When `pairs_confirmed=True`, the variant's `pairs` field is
    # the trusted ground truth derived from `segments`. When False,
    # the user has not yet reviewed step 1 and step 2 is hidden.
    segments: list[SegmentSpec] = field(default_factory=list)
    pairs_confirmed: bool = False
    # Cached raw trace (time_min, abs) loaded lazily when the variant
    # tab is first opened. Declared here so callers can safely read
    # `s._raw_trace is None` without needing the FileTab to have been
    # built — important for reextract_all and other App-level paths
    # that touch every session regardless of tab state.
    _raw_trace: Optional[tuple] = field(default=None, repr=False, compare=False)


@dataclass
class AppState:
    folder: Optional[Path] = None
    sessions: list[FileSession] = field(default_factory=list)
    params: PipelineParams = field(default_factory=PipelineParams)


# ---------------------------------------------------------------------------
# Helpers for slope extraction
# ---------------------------------------------------------------------------

def auto_detect_segments(path: Path, params: Optional[PipelineParams] = None
                          ) -> tuple[list[SegmentSpec], Optional[str]]:
    """Run the segmenter + classifier on a CSV and return a list of
    SegmentSpec (with auto labels). Does NOT pair them or fit slopes —
    that is done by `build_pairs_from_segments`. Returning segments and
    pairs separately lets the user manually edit the segment list (in
    step 1) before pairs are derived."""
    if params is None:
        params = PipelineParams()
    try:
        df = kp.load_cary_csv(path)
    except Exception as exc:
        return [], f"Could not parse Cary CSV: {exc}"

    t = df["time_min"].to_numpy()
    a = df["abs"].to_numpy()
    try:
        segments = kp.detect_segments(t, a, min_seg_sec=params.min_seg_sec)
        if not segments:
            return [], "No segments detected."
        labels = kp.classify_segments(segments, t, a, force_alternating=False)
        # Refine enzyme starts to the post-spike absorbance minimum
        refined: list[SegmentSpec] = []
        for (s, e), lab in zip(segments, labels):
            if lab == "enzyme" and params.extend_to_minimum:
                s = kp._extend_start_to_minimum(
                    t, a, s,
                    max_lookback_sec=params.extend_lookback_sec,
                    smooth_thresh=params.extend_smooth_thresh,
                )
            refined.append(SegmentSpec(s_idx=int(s), e_idx=int(e),
                                       label=lab))
    except Exception as exc:
        return [], f"Segmentation failed: {exc}"

    return refined, None


def build_pairs_from_segments(
    path: Path,
    segments: list[SegmentSpec],
    params: Optional[PipelineParams] = None,
) -> tuple[list[PairRow], Optional[str]]:
    """Take a (possibly hand-edited) list of SegmentSpec and produce
    PairRow objects with fitted slopes. Pairs are formed by matching
    each enzyme segment with the most recently-preceding blank segment.

    The segments must be in chronological order (sorted by `s_idx`).
    Both blanks and enzymes get fitted; the blank fit skips post-spike
    transients via `_refit_blank_segment`; the enzyme fit uses the
    auto-window-finder in `kp.fit_enzyme_initial_rate`.
    """
    if params is None:
        params = PipelineParams()
    if not segments:
        return [], None
    try:
        df = kp.load_cary_csv(path)
    except Exception as exc:
        return [], f"Could not parse Cary CSV: {exc}"
    t = df["time_min"].to_numpy()
    a = df["abs"].to_numpy()

    # Sort defensively
    segs_sorted = sorted(segments, key=lambda sp: sp.s_idx)

    # Fit each segment
    fits: list[kp.LinearFit] = []
    for sp in segs_sorted:
        ts = t[sp.s_idx:sp.e_idx]
        ys = a[sp.s_idx:sp.e_idx]
        if len(ts) < 5:
            fits.append(kp.LinearFit(
                slope=float("nan"), intercept=float("nan"),
                r2=float("nan"), n_points=len(ts),
                t_start=float(ts[0]) if len(ts) else 0.0,
                t_end=float(ts[-1]) if len(ts) else 0.0,
                window_sec=0.0, meets_r2_threshold=False))
            continue
        if sp.label == "blank":
            fits.append(kp.fit_blank(
                ts, ys, head_trim_sec=params.head_trim_sec,
                tail_trim_sec=params.tail_trim_sec))
        else:
            if params.fit_method == "curvature":
                fits.append(kp.fit_enzyme_linear_robust(
                    ts, ys,
                    head_trim_sec=params.head_trim_sec,
                    tail_trim_sec=params.tail_trim_sec,
                    min_window_sec=params.min_window_sec,
                    max_slope_dev=params.max_slope_dev,
                    curvature_sig_sigma=params.curvature_sig_sigma,
                ))
            else:
                fits.append(kp.fit_enzyme_initial_rate(
                    ts, ys,
                    head_trim_sec=params.head_trim_sec,
                    tail_trim_sec=params.tail_trim_sec,
                    r2_thresh=params.r2_thresh,
                    slope_tol=params.slope_tol,
                    min_window_sec=params.min_window_sec,
                ))

    # Pair every enzyme with most recently-preceding blank
    rows: list[PairRow] = []
    pair_no = 0
    last_blank_i: Optional[int] = None
    skipped_enzymes = 0
    for i, sp in enumerate(segs_sorted):
        if sp.label == "blank":
            last_blank_i = i
            continue
        if sp.label != "enzyme":
            continue
        if last_blank_i is None:
            skipped_enzymes += 1
            continue
        pair_no += 1
        fe = fits[i]
        sb_seg = segs_sorted[last_blank_i]
        # Override the auto blank fit with the transient-aware version
        slope_blank, intercept_blank, sb_fit, eb_fit = _refit_blank_segment(
            t, a, sb_seg.s_idx, sb_seg.e_idx,
            head_trim_sec=params.head_trim_sec,
            tail_trim_sec=params.tail_trim_sec,
        )
        row = PairRow(
            pair=pair_no,
            t_enzyme_start=fe.t_start,
            slope_blank=slope_blank,
            slope_enzyme=fe.slope,
            slope_net=fe.slope - slope_blank,
            r2_enzyme=fe.r2,
            meets_r2=getattr(fe, "meets_r2_threshold", True),
            window_enzyme_s=fe.window_sec,
            _enzyme_seg=(sp.s_idx, sp.e_idx),
            _blank_seg=(sb_seg.s_idx, sb_seg.e_idx),
            _blank_fit_seg=(sb_fit, eb_fit),
            _blank_intercept=intercept_blank,
        )
        row._auto_t_enzyme_start = row.t_enzyme_start
        row._auto_window_enzyme_s = row.window_enzyme_s
        row._auto_slope_enzyme = row.slope_enzyme
        row._auto_r2_enzyme = row.r2_enzyme
        row._auto_slope_net = row.slope_net
        row._auto_meets_r2 = row.meets_r2
        rows.append(row)

    if len(rows) == len(DEFAULT_CONCENTRATIONS_UM):
        for r, c in zip(rows, DEFAULT_CONCENTRATIONS_UM):
            r.concentration_uM = c

    err_msg = None
    if skipped_enzymes > 0:
        err_msg = (f"{skipped_enzymes} enzyme segment(s) had no preceding "
                   "blank and were dropped.")
    return rows, err_msg


def extract_pairs(path: Path, params: Optional[PipelineParams] = None
                  ) -> tuple[list[PairRow], Optional[str]]:
    """Convenience wrapper: auto-detect + build pairs in one call.
    Kept for backward compatibility with code paths that don't need
    the segment list."""
    segments, err = auto_detect_segments(path, params=params)
    if err is not None and not segments:
        return [], err
    rows, err2 = build_pairs_from_segments(path, segments, params=params)
    return rows, err or err2


def _find_blank_settled_start(
    t: np.ndarray, a: np.ndarray,
    window_sec: float = 1.5,
    settle_thresh_AU_per_min: float = 0.05,
    max_skip_sec: float = 5.0,
) -> int:
    """Return the index in (t, a) where the blank trace has settled
    after any post-spike transient.

    The dilation buffer in `detect_segments` removes the spike itself,
    but on the blank side a residual rise/ringing (a few hundred ms to
    a couple of seconds) often remains. A linear fit including this
    head fragment over-estimates the substrate decomposition slope by
    an order of magnitude.

    Strategy: compute the rolling forward slope (AU/min) over a short
    window. The blank is "settled" once the slope drops below
    `settle_thresh_AU_per_min`. Skip up to `max_skip_sec` of head data;
    if no settled point is found within that budget, return 0 (don't
    skip anything — the segment is genuinely sloping and the fit
    should reflect that).
    """
    n = len(t)
    if n < 10:
        return 0
    dt_min = float(np.mean(np.diff(t)))
    if dt_min <= 0:
        return 0
    win = max(5, int((window_sec / 60.0) / dt_min))
    win = min(win, n // 3)
    max_skip_pts = int(max_skip_sec / 60.0 / dt_min)

    for i in range(min(max_skip_pts, n - win)):
        m, _, _ = _linreg_simple(t[i:i + win], a[i:i + win])
        if abs(m) <= settle_thresh_AU_per_min:
            return i
    return 0


def _linreg_simple(x: np.ndarray, y: np.ndarray) -> tuple[float, float, float]:
    """Lightweight linear regression returning (slope, intercept, r2)."""
    if len(x) < 3:
        return float("nan"), float("nan"), float("nan")
    m, b = np.polyfit(x, y, 1)
    yhat = m * x + b
    ss_res = float(np.sum((y - yhat) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else 1.0
    return float(m), float(b), float(r2)


def _refit_blank_segment(t: np.ndarray, a: np.ndarray,
                          s: int, e: int,
                          head_trim_sec: float, tail_trim_sec: float,
                          ) -> tuple[float, float, int, int]:
    """Refit a blank segment after skipping any post-spike transient.

    Returns (slope_AU_per_min, intercept, fit_start_idx, fit_end_idx)
    where the fit indices are absolute (into `t`/`a`)."""
    skip = _find_blank_settled_start(t[s:e], a[s:e])
    s_eff = s + skip
    # Apply user-set trims on top of the settle skip
    dt_min = float(np.mean(np.diff(t[s:e]))) if (e - s) > 1 else 0.0
    head_pts = int(head_trim_sec / 60.0 / dt_min) if dt_min > 0 else 0
    tail_pts = int(tail_trim_sec / 60.0 / dt_min) if dt_min > 0 else 0
    s_eff = min(s_eff + head_pts, e - 5)
    e_eff = max(e - tail_pts, s_eff + 5)
    x_fit, y_fit = t[s_eff:e_eff], a[s_eff:e_eff]
    if len(x_fit) < 5:
        # Fall back to whatever we have
        x_fit, y_fit = t[s:e], a[s:e]
        s_eff, e_eff = s, e
    m, b, _ = _linreg_simple(x_fit, y_fit)
    return m, b, s_eff, e_eff


def refit_pair_window(p: PairRow, t_full: np.ndarray, a_full: np.ndarray,
                      new_t_start: float, new_t_end: float,
                      r2_thresh: float = 0.995) -> bool:
    """Recompute slope/R^2 for an enzyme pair over a user-chosen window.

    Returns True if the refit succeeded. The blank slope is left
    untouched (we keep the auto-extracted blank baseline). The pair is
    flagged as `manually_adjusted = True`.
    """
    if p._enzyme_seg is None:
        return False
    s_e, e_e = p._enzyme_seg
    xs = t_full[s_e:e_e]
    ys = a_full[s_e:e_e]
    if len(xs) < 5:
        return False

    # Clamp the window to the segment bounds.
    new_t_start = max(float(xs[0]), float(new_t_start))
    new_t_end = min(float(xs[-1]), float(new_t_end))
    if new_t_end - new_t_start < (1.0 / 60.0):  # require at least 1 second
        return False

    mask = (xs >= new_t_start) & (xs <= new_t_end)
    if mask.sum() < 5:
        return False
    x_fit, y_fit = xs[mask], ys[mask]
    m, b = np.polyfit(x_fit, y_fit, 1)
    yhat = m * x_fit + b
    ss_res = float(np.sum((y_fit - yhat) ** 2))
    ss_tot = float(np.sum((y_fit - y_fit.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")

    p.t_enzyme_start = float(x_fit[0])
    p.window_enzyme_s = float((x_fit[-1] - x_fit[0]) * 60.0)
    p.slope_enzyme = float(m)
    p.r2_enzyme = float(r2)
    p.slope_net = float(m - p.slope_blank)
    p.meets_r2 = bool(r2 >= r2_thresh)
    p.manually_adjusted = True
    return True


def revert_pair_to_auto(p: PairRow) -> None:
    """Undo a manual window adjustment; restore the auto-fit values."""
    if not p.manually_adjusted:
        return
    p.t_enzyme_start = p._auto_t_enzyme_start
    p.window_enzyme_s = p._auto_window_enzyme_s
    p.slope_enzyme = p._auto_slope_enzyme
    p.r2_enzyme = p._auto_r2_enzyme
    p.slope_net = p._auto_slope_net
    p.meets_r2 = p._auto_meets_r2
    p.manually_adjusted = False


SIDECAR_FILENAME = ".kinetics_session.json"
SIDECAR_VERSION = 2

# Per-script user defaults file. Lives next to kinetics_gui.py so that
# a user's preferred parameter values follow the installation, not the
# data folder. Saved/restored via the "Save as default" /
# "Restore factory defaults" buttons on the Parameters tab.
USER_DEFAULTS_FILENAME = "kinetics_defaults.json"


def _user_defaults_path() -> Path:
    """Return the path where user defaults are stored. Lives next to
    `kinetics_gui.py` itself."""
    return Path(__file__).resolve().parent / USER_DEFAULTS_FILENAME


def load_user_defaults() -> Optional[dict]:
    """Read the user-defaults JSON if present. Returns the parsed dict
    on success, None if missing or malformed."""
    import json
    p = _user_defaults_path()
    if not p.is_file():
        return None
    try:
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else None
    except Exception:
        return None


def save_user_defaults(params: "PipelineParams") -> tuple[bool, str]:
    """Write the given PipelineParams to the user-defaults JSON file.

    Returns (ok, message). The message is the resolved file path on
    success, or an error description on failure."""
    import json
    p = _user_defaults_path()
    try:
        data = asdict(params)
        tmp = p.with_suffix(p.suffix + ".tmp")
        with tmp.open("w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        tmp.replace(p)
        return True, str(p)
    except Exception as exc:
        return False, str(exc)


def clear_user_defaults() -> tuple[bool, str]:
    """Delete the user-defaults JSON file if it exists. Returns
    (ok, message). Removing a missing file counts as success."""
    p = _user_defaults_path()
    try:
        if p.is_file():
            p.unlink()
        return True, str(p)
    except Exception as exc:
        return False, str(exc)


def make_params_with_user_defaults() -> "PipelineParams":
    """Build a `PipelineParams` instance, applying user-saved defaults
    on top of the hardcoded ones. If no user defaults exist or any
    field is malformed, the hardcoded defaults are used."""
    p = PipelineParams()
    saved = load_user_defaults()
    if not saved:
        return p
    # Apply only known fields with type-compatible values; ignore the rest
    for attr in p.__dataclass_fields__:
        if attr not in saved:
            continue
        cur = getattr(p, attr)
        new = saved[attr]
        # Coerce numeric values to the right type when possible
        try:
            if isinstance(cur, bool):
                new = bool(new)
            elif isinstance(cur, int) and not isinstance(cur, bool):
                new = int(new)
            elif isinstance(cur, float):
                new = float(new)
            elif isinstance(cur, str):
                new = str(new)
        except Exception:
            continue  # malformed entry, skip
        setattr(p, attr, new)
    return p

# v1 → v2 migration: v1 stored only the inputs (segments, overrides).
# Reload at the time was reconstructive — slopes were recomputed from
# segments + current parameters, which meant the displayed value could
# differ from the value confirmed because of a parameter change between
# sessions. v2 stores the COMPUTED slopes too (a "checkpoint"), so
# reload is deterministic: you see exactly what you saved. Re-extract
# explicitly overwrites with fresh computed values.
#
# When reading a v1 sidecar, the loader recomputes slopes from segments
# at current parameters (v1 behaviour) and shows a one-time warning.


def load_sidecar(folder: Path) -> dict:
    """Read the per-folder sidecar if present. Returns an empty dict
    on any error (no sidecar yet, malformed, etc.)."""
    import json
    p = folder / SIDECAR_FILENAME
    if not p.is_file():
        return {}
    try:
        with p.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_sidecar(folder: Path, sessions: list["FileSession"]) -> None:
    """Write the sidecar JSON. Sidecar v2 captures every value needed
    to reconstruct the session deterministically: the segments (inputs)
    AND the computed slopes (checkpoint). Reload reads slopes verbatim;
    re-extract regenerates them at current parameters."""
    import json
    data = {"version": SIDECAR_VERSION, "files": {}}
    for s in sessions:
        pair_snapshots = []
        for p in s.pairs:
            if p.manually_adjusted:
                manual_window = [float(p.t_enzyme_start),
                                 float(p.t_enzyme_start
                                       + p.window_enzyme_s / 60.0)]
            else:
                manual_window = None
            pair_snapshots.append({
                # Inputs the user can edit
                "include": bool(p.include),
                "concentration_uM": (None
                                     if not np.isfinite(p.concentration_uM)
                                     else float(p.concentration_uM)),
                "manual_window": manual_window,
                # Computed values (checkpoint) — used verbatim on reload
                "pair_no": int(p.pair),
                "t_enzyme_start": float(p.t_enzyme_start),
                "window_enzyme_s": float(p.window_enzyme_s),
                "slope_enzyme": float(p.slope_enzyme),
                "slope_blank": float(p.slope_blank),
                "slope_net": float(p.slope_net),
                "r2_enzyme": float(p.r2_enzyme),
                "meets_r2": bool(p.meets_r2),
                "manually_adjusted": bool(p.manually_adjusted),
                "enzyme_seg": (list(p._enzyme_seg)
                               if p._enzyme_seg else None),
                "blank_seg": (list(p._blank_seg)
                              if p._blank_seg else None),
                "blank_fit_seg": (list(p._blank_fit_seg)
                                  if p._blank_fit_seg else None),
                "blank_intercept": (float(p._blank_intercept)
                                    if np.isfinite(p._blank_intercept)
                                    else None),
                # Auto-fit snapshots — let manual edits revert correctly
                "auto_t_enzyme_start": float(p._auto_t_enzyme_start),
                "auto_window_enzyme_s": float(p._auto_window_enzyme_s),
                "auto_slope_enzyme": float(p._auto_slope_enzyme),
                "auto_r2_enzyme": float(p._auto_r2_enzyme),
                "auto_slope_net": float(p._auto_slope_net),
                "auto_meets_r2": bool(p._auto_meets_r2),
            })
        data["files"][s.path.name] = {
            "segments": [seg.to_dict() for seg in s.segments],
            "pairs_confirmed": bool(s.pairs_confirmed),
            "enzyme_conc_uM": float(s.enzyme_conc_uM),
            "epsilon_M1cm1": float(s.epsilon_M1cm1),
            "pair_snapshots": pair_snapshots,
        }
    p = folder / SIDECAR_FILENAME
    try:
        # Write atomically: write to a tmp file then rename. Avoids
        # leaving a half-written sidecar if the process is killed.
        tmp = p.with_suffix(p.suffix + ".tmp")
        with tmp.open("w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        tmp.replace(p)
    except Exception:
        pass


def restore_pairs_from_snapshots(snapshots: list[dict]
                                  ) -> list["PairRow"]:
    """Reconstruct PairRow objects from sidecar v2 snapshots, applying
    saved values verbatim. No recomputation."""
    pairs: list[PairRow] = []
    for snap in snapshots:
        p = PairRow(
            pair=int(snap.get("pair_no", len(pairs) + 1)),
            t_enzyme_start=float(snap["t_enzyme_start"]),
            slope_blank=float(snap["slope_blank"]),
            slope_enzyme=float(snap["slope_enzyme"]),
            slope_net=float(snap["slope_net"]),
            r2_enzyme=float(snap["r2_enzyme"]),
            meets_r2=bool(snap["meets_r2"]),
            window_enzyme_s=float(snap["window_enzyme_s"]),
            concentration_uM=(float(snap["concentration_uM"])
                              if snap.get("concentration_uM") is not None
                              else float("nan")),
            include=bool(snap.get("include", True)),
            manually_adjusted=bool(snap.get("manually_adjusted", False)),
            _auto_t_enzyme_start=float(snap.get("auto_t_enzyme_start",
                                                snap["t_enzyme_start"])),
            _auto_window_enzyme_s=float(snap.get("auto_window_enzyme_s",
                                                  snap["window_enzyme_s"])),
            _auto_slope_enzyme=float(snap.get("auto_slope_enzyme",
                                              snap["slope_enzyme"])),
            _auto_r2_enzyme=float(snap.get("auto_r2_enzyme",
                                           snap["r2_enzyme"])),
            _auto_slope_net=float(snap.get("auto_slope_net",
                                           snap["slope_net"])),
            _auto_meets_r2=bool(snap.get("auto_meets_r2",
                                         snap["meets_r2"])),
            _enzyme_seg=(tuple(snap["enzyme_seg"])
                         if snap.get("enzyme_seg") else None),
            _blank_seg=(tuple(snap["blank_seg"])
                        if snap.get("blank_seg") else None),
            _blank_fit_seg=(tuple(snap["blank_fit_seg"])
                            if snap.get("blank_fit_seg") else None),
            _blank_intercept=(float(snap["blank_intercept"])
                              if snap.get("blank_intercept") is not None
                              else float("nan")),
        )
        pairs.append(p)
    return pairs


def discover_csv_files(folder: Path) -> list[Path]:
    """Return all *.csv files in the folder, excluding obvious pipeline
    outputs (anything ending in _kinetics.csv) and dotfiles."""
    out = []
    for p in sorted(folder.glob("*.csv")):
        if p.name.startswith("."):
            continue
        if p.name.endswith("_kinetics.csv"):
            continue
        out.append(p)
    return out


# ---------------------------------------------------------------------------
# Recompute
# ---------------------------------------------------------------------------

def recompute_session(s: FileSession,
                       params: Optional[PipelineParams] = None) -> None:
    """Recompute kobs and kcat/KM for a session in place.

    The `params` argument decides which kcat-fitting method is used:
    'weighted', 'per_replicate', or 'bootstrap'. If `params` is None,
    falls back to defaults — useful for tests and standalone callers.
    """
    if params is None:
        params = PipelineParams()
    conc, kobs, order = [], [], []
    for r in s.pairs:
        if not r.include:
            continue
        if not np.isfinite(r.concentration_uM) or r.concentration_uM <= 0:
            continue
        v = velocity_uM_per_s(r.slope_net, s.epsilon_M1cm1)
        k = kobs_per_s(v, s.enzyme_conc_uM)
        if not np.isfinite(k):
            continue
        conc.append(r.concentration_uM)
        kobs.append(k)
        order.append(r.pair)

    method = params.kcat_method
    # Reset all uncertainty fields so a method switch doesn't leave
    # stale values in place
    s.kcat_over_km_se = float("nan")
    s.kcat_ci_lo = float("nan")
    s.kcat_ci_hi = float("nan")
    s.kcat_method_used = method

    if method == "per_replicate":
        slope, sd, intercept, r2, n_reps, n_pts = (
            fit_kcat_over_km_per_replicate(conc, kobs, order))
        s.kcat_over_km = slope
        s.kcat_over_km_se = sd  # for the per-replicate method this is the
                                # SD across replicate slopes, not a
                                # regression SE — labeled accordingly
                                # by the UI/exporter
        s.fit_intercept = intercept
        s.fit_r2 = r2
        s.n_used = n_pts
    elif method == "bootstrap":
        n_iter = int(getattr(params, "bootstrap_iter", 10000))
        slope, ci_lo, ci_hi, intercept, r2, n_reps, n_pts = (
            fit_kcat_over_km_bootstrap(conc, kobs, order, n_iter=n_iter))
        s.kcat_over_km = slope
        s.kcat_ci_lo = ci_lo
        s.kcat_ci_hi = ci_hi
        s.fit_intercept = intercept
        s.fit_r2 = r2
        s.n_used = n_pts
    else:  # 'weighted' (default)
        slope, slope_se, intercept, r2, n = (
            fit_kcat_over_km_weighted(conc, kobs))
        s.kcat_over_km = slope
        s.kcat_over_km_se = slope_se
        s.fit_intercept = intercept
        s.fit_r2 = r2
        s.n_used = n
        s.kcat_method_used = "weighted"


# ---------------------------------------------------------------------------
# Tk helpers
# ---------------------------------------------------------------------------

class ScrollableFrame(ttk.Frame):
    """A vertically-scrollable container."""
    def __init__(self, master, **kw):
        super().__init__(master, **kw)
        canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        vsb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self.inner = ttk.Frame(canvas)
        self._win = canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfigure(self._win, width=e.width))
        # Mouse-wheel scroll
        def _on_wheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_wheel)


def parse_float(s: str) -> float:
    try:
        return float(s.replace(",", "."))
    except (ValueError, AttributeError):
        return float("nan")


def fmt(x: float, n: int = 4) -> str:
    if x is None or not np.isfinite(x):
        return ""
    if abs(x) >= 1e4 or (x != 0 and abs(x) < 1e-3):
        return f"{x:.{n}e}"
    return f"{x:.{n}g}"


# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------

class SetupTab(ttk.Frame):
    """Editable per-enzyme table: variant | [E] | epsilon | n | n_excl |
    kcat/KM | R2 | intercept."""

    HEADERS = ["#", "Variant", "[E] (µM)", "ε (M⁻¹cm⁻¹)",
               "n pairs", "n excl", "kcat/KM (M⁻¹s⁻¹)",
               "R² (fit)", "intercept (s⁻¹)"]

    def __init__(self, master, app: "App"):
        super().__init__(master)
        self.app = app
        self._row_widgets: list[dict] = []

        info = ttk.Label(self, text=(
            "Edit [E] and ε per variant. Defaults: ε = "
            f"{DEFAULT_EPSILON_M1CM1:g} M⁻¹cm⁻¹, [E] = "
            f"{DEFAULT_ENZYME_CONC_UM} µM (placeholder — please update). "
            "Tab/Enter to commit. Click 'Apply ε to all' to broadcast."),
            wraplength=1000, justify="left")
        info.pack(fill="x", padx=8, pady=(8, 4))

        # Quick-action row
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=8, pady=4)
        ttk.Label(bar, text="Set ε for all rows:").pack(side="left")
        self._eps_entry = ttk.Entry(bar, width=14)
        self._eps_entry.insert(0, f"{DEFAULT_EPSILON_M1CM1:g}")
        self._eps_entry.pack(side="left", padx=4)
        ttk.Button(bar, text="Apply ε to all",
                   command=self._apply_epsilon_all).pack(side="left", padx=4)
        ttk.Label(bar, text="   Set [E] for all rows:").pack(side="left")
        self._ec_entry = ttk.Entry(bar, width=8)
        self._ec_entry.pack(side="left", padx=4)
        ttk.Button(bar, text="Apply [E] to all",
                   command=self._apply_ec_all).pack(side="left", padx=4)

        # Scrollable table
        scroll = ScrollableFrame(self)
        scroll.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        self._table = scroll.inner

        # Header row
        for j, h in enumerate(self.HEADERS):
            ttk.Label(self._table, text=h, font=("TkDefaultFont", 9, "bold")
                      ).grid(row=0, column=j, padx=4, pady=2, sticky="w")

    def rebuild(self):
        # Clear existing data rows (keep header at row 0). Some entries
        # in the row-widget dict are tk Variables (StringVar etc.) which
        # have no .destroy(); only call .destroy() on real widgets.
        for w in self._row_widgets:
            for v in w.values():
                if hasattr(v, "destroy"):
                    v.destroy()
        self._row_widgets.clear()

        for i, s in enumerate(self.app.state.sessions, start=1):
            row = i
            widgets = {}
            ttk.Label(self._table, text=str(i)).grid(
                row=row, column=0, padx=4, pady=1, sticky="w")
            widgets["idx"] = self._table.grid_slaves(row=row, column=0)[0]

            name_lbl = ttk.Label(self._table, text=s.name)
            name_lbl.grid(row=row, column=1, padx=4, pady=1, sticky="w")
            widgets["name"] = name_lbl

            ec_var = tk.StringVar(value=fmt(s.enzyme_conc_uM, 6))
            ec_entry = ttk.Entry(self._table, width=10, textvariable=ec_var)
            ec_entry.grid(row=row, column=2, padx=4, pady=1, sticky="w")
            ec_entry.bind("<FocusOut>",
                          lambda _e, sess=s, var=ec_var: self._on_ec(sess, var))
            ec_entry.bind("<Return>",
                          lambda _e, sess=s, var=ec_var: self._on_ec(sess, var))
            widgets["ec"] = ec_entry
            widgets["ec_var"] = ec_var

            eps_var = tk.StringVar(value=fmt(s.epsilon_M1cm1, 8))
            eps_entry = ttk.Entry(self._table, width=14, textvariable=eps_var)
            eps_entry.grid(row=row, column=3, padx=4, pady=1, sticky="w")
            eps_entry.bind("<FocusOut>",
                           lambda _e, sess=s, var=eps_var: self._on_eps(sess, var))
            eps_entry.bind("<Return>",
                           lambda _e, sess=s, var=eps_var: self._on_eps(sess, var))
            widgets["eps"] = eps_entry
            widgets["eps_var"] = eps_var

            n_pairs = ttk.Label(self._table, text=str(len(s.pairs)))
            n_pairs.grid(row=row, column=4, padx=4, pady=1, sticky="w")
            widgets["n_pairs"] = n_pairs

            n_excl = ttk.Label(self._table, text="0")
            n_excl.grid(row=row, column=5, padx=4, pady=1, sticky="w")
            widgets["n_excl"] = n_excl

            kk = ttk.Label(self._table, text="")
            kk.grid(row=row, column=6, padx=4, pady=1, sticky="w")
            widgets["kk"] = kk

            r2 = ttk.Label(self._table, text="")
            r2.grid(row=row, column=7, padx=4, pady=1, sticky="w")
            widgets["r2"] = r2

            ic = ttk.Label(self._table, text="")
            ic.grid(row=row, column=8, padx=4, pady=1, sticky="w")
            widgets["ic"] = ic

            self._row_widgets.append(widgets)

        self.refresh_results()

    def refresh_results(self):
        """Pull computed values from each session and update labels."""
        for s, w in zip(self.app.state.sessions, self._row_widgets):
            n_excl = sum(1 for p in s.pairs if not p.include)
            w["n_excl"].configure(text=str(n_excl))
            w["kk"].configure(text=fmt(s.kcat_over_km, 4))
            w["r2"].configure(text=fmt(s.fit_r2, 4))
            w["ic"].configure(text=fmt(s.fit_intercept, 4))

    def _on_ec(self, sess: FileSession, var: tk.StringVar):
        v = parse_float(var.get())
        if not np.isfinite(v) or v <= 0:
            messagebox.showwarning(
                "Invalid value",
                f"[E] must be a positive number; got '{var.get()}'.")
            var.set(fmt(sess.enzyme_conc_uM, 6))
            return
        sess.enzyme_conc_uM = v
        recompute_session(sess, self.app.state.params)
        self.app.recomputed_one(sess)

    def _on_eps(self, sess: FileSession, var: tk.StringVar):
        v = parse_float(var.get())
        if not np.isfinite(v) or v <= 0:
            messagebox.showwarning(
                "Invalid value",
                f"ε must be a positive number; got '{var.get()}'.")
            var.set(fmt(sess.epsilon_M1cm1, 8))
            return
        sess.epsilon_M1cm1 = v
        recompute_session(sess, self.app.state.params)
        self.app.recomputed_one(sess)

    def _apply_epsilon_all(self):
        v = parse_float(self._eps_entry.get())
        if not np.isfinite(v) or v <= 0:
            messagebox.showwarning("Invalid value", "ε must be positive.")
            return
        for s, w in zip(self.app.state.sessions, self._row_widgets):
            s.epsilon_M1cm1 = v
            w["eps_var"].set(fmt(v, 8))
            recompute_session(s, self.app.state.params)
        self.app.recomputed_all()

    def _apply_ec_all(self):
        v = parse_float(self._ec_entry.get())
        if not np.isfinite(v) or v <= 0:
            messagebox.showwarning("Invalid value", "[E] must be positive.")
            return
        for s, w in zip(self.app.state.sessions, self._row_widgets):
            s.enzyme_conc_uM = v
            w["ec_var"].set(fmt(v, 6))
            recompute_session(s, self.app.state.params)
        self.app.recomputed_all()


class ParametersTab(ttk.Frame):
    """Edit the slope-extraction pipeline parameters and re-extract."""

    # Each entry: (attr, label, tooltip, type, validator)
    # type can be: float, bool, or a tuple ('choice', [option1, option2, ...])
    FIELDS = [
        ("kcat_method", "kcat/KM combining method",
         "How to combine per-pair k_obs into one k_cat/K_M per variant. "
         "'weighted': one regression using per-[S] replicate SDs as "
         "weights, reports a slope SE. "
         "'per_replicate': fit each technical replicate separately, "
         "report mean ± SD across replicates. "
         "'bootstrap': fit on the full data, report 95% CI from "
         "replicate-clustered bootstrap.",
         ("choice", ["weighted", "per_replicate", "bootstrap"]), None),
        ("bootstrap_iter", "Bootstrap iterations",
         "Number of resamples for the bootstrap method. Default 10000 "
         "is fine for most assays.",
         float, lambda v: v >= 100),
        ("fit_method", "Enzyme fit method",
         "How to choose the linear region of each enzyme segment. "
         "'curvature' uses a noise-robust quadratic test (recommended); "
         "'r2_window' uses the older R²-based window growth.",
         ("choice", ["curvature", "r2_window"]), None),
        ("max_slope_dev", "Max slope deviation (curvature method)",
         "Allowed fractional slope drop across the fit window before "
         "the curvature method truncates. 0.05 = 5% (default).",
         float, lambda v: 0 < v < 1),
        ("curvature_sig_sigma", "Curvature significance σ (curvature method)",
         "Minimum statistical significance of the quadratic curvature "
         "term before the fit window is truncated. Higher = more "
         "tolerant of small curvatures.",
         float, lambda v: v >= 1),
        ("min_seg_sec",          "Min segment length (s)",
         "Discard segments shorter than this; spike settle stubs.",
         float, lambda v: v >= 5),
        ("min_window_sec",       "Min fit window (s)",
         "Smallest window the auto-fitter will consider. Going below "
         "about 0.5 s on 10 Hz data gives 5 or fewer points per fit "
         "and noisy slope estimates; 0.3 s is the hard floor.",
         float, lambda v: v >= 0.3),
        ("r2_thresh",            "R² threshold (r2_window method)",
         "Minimum R² for the auto-fit window to be accepted. "
         "Only applies when fit_method = 'r2_window'.",
         float, lambda v: 0 < v <= 1),
        ("slope_tol",            "Slope tolerance (r2_window method)",
         "Reject windows whose slope drops below "
         "(1 − tol) × initial-velocity reference. Higher = more lenient.",
         float, lambda v: 0 <= v <= 1),
        ("head_trim_sec",        "Head trim (s)",
         "Extra trim from the start of each segment after the dilation "
         "buffer.", float, lambda v: v >= 0),
        ("tail_trim_sec",        "Tail trim (s)",
         "Trim from the end of each segment.",
         float, lambda v: v >= 0),
        ("extend_to_minimum",    "Extend to A-minimum",
         "Walk back from the dilation-buffered enzyme segment start to "
         "the post-spike absorbance minimum.",
         bool, None),
        ("extend_lookback_sec",  "Extend lookback (s)",
         "Max seconds to walk back when searching for the absorbance "
         "minimum.", float, lambda v: v >= 0),
        ("extend_smooth_thresh", "Extend smooth thresh (|ΔA|)",
         "Point-to-point |ΔA| above this is treated as a spike edge "
         "during the backward walk.", float, lambda v: v >= 0),
    ]

    def __init__(self, master, app: "App"):
        super().__init__(master)
        self.app = app
        self._vars: dict[str, tk.Variable] = {}

        info = ttk.Label(self, text=(
            "Pipeline parameters used during slope extraction. Editing a "
            "value here does NOT reanalyse anything until you click "
            "'Re-extract slopes'. Re-extraction wipes any manual fit-"
            "window adjustments you made on the variant tabs."),
            wraplength=900, justify="left")
        info.pack(fill="x", padx=8, pady=(8, 4))

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=8, pady=4)

        # Two-column layout: label / entry / description
        for row, (attr, label, tooltip, kind, _validator) in enumerate(self.FIELDS):
            ttk.Label(body, text=label).grid(row=row, column=0,
                                             padx=6, pady=3, sticky="w")
            cur = getattr(self.app.state.params, attr)
            if kind is bool:
                var = tk.BooleanVar(value=bool(cur))
                ttk.Checkbutton(body, variable=var).grid(
                    row=row, column=1, padx=6, pady=3, sticky="w")
            elif isinstance(kind, tuple) and kind[0] == "choice":
                var = tk.StringVar(value=str(cur))
                cb = ttk.Combobox(body, width=14, textvariable=var,
                                  values=kind[1], state="readonly")
                cb.grid(row=row, column=1, padx=6, pady=3, sticky="w")
            else:
                var = tk.StringVar(value=fmt(cur, 6))
                ttk.Entry(body, width=12, textvariable=var).grid(
                    row=row, column=1, padx=6, pady=3, sticky="w")
            self._vars[attr] = var
            ttk.Label(body, text=tooltip, foreground="grey",
                      wraplength=600, justify="left").grid(
                row=row, column=2, padx=6, pady=3, sticky="w")

        # Action row
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=8, pady=(8, 8))
        ttk.Button(bar, text="Restore defaults",
                   command=self._restore_defaults).pack(side="left")
        ttk.Button(bar, text="Save as default",
                   command=self._save_as_default).pack(side="left",
                                                        padx=(8, 0))
        ttk.Button(bar, text="Clear saved default",
                   command=self._clear_saved_default).pack(side="left",
                                                            padx=(8, 0))
        ttk.Button(bar, text="Re-extract slopes",
                   command=self._reextract).pack(side="right")

    def _read_into_params(self) -> tuple[bool, str]:
        """Validate fields and write into self.app.state.params. Returns
        (ok, error_message)."""
        new_params = PipelineParams(**asdict(self.app.state.params))
        for attr, _label, _tt, kind, validator in self.FIELDS:
            var = self._vars[attr]
            if kind is bool:
                v = bool(var.get())
            elif isinstance(kind, tuple) and kind[0] == "choice":
                v = str(var.get())
                if v not in kind[1]:
                    return False, f"'{_label}' must be one of {kind[1]}, got {v!r}"
            else:
                raw = var.get()
                v = parse_float(raw)
                if not np.isfinite(v):
                    return False, f"'{_label}' is not a valid number: {raw!r}"
                if validator and not validator(v):
                    return False, f"'{_label}' value {v} is out of range."
                # Integer-valued params get cast at read time so the
                # dataclass holds an int, not a float-via-float.
                if attr in ("bootstrap_iter",):
                    v = int(v)
            setattr(new_params, attr, v)
        self.app.state.params = new_params
        return True, ""

    def _restore_defaults(self):
        defaults = PipelineParams()
        for attr, _label, _tt, kind, _v in self.FIELDS:
            cur = getattr(defaults, attr)
            if kind is bool:
                self._vars[attr].set(bool(cur))
            elif isinstance(kind, tuple) and kind[0] == "choice":
                self._vars[attr].set(str(cur))
            else:
                self._vars[attr].set(fmt(cur, 6))

    def _save_as_default(self):
        """Persist the current Parameters-tab values to the
        kinetics_defaults.json file next to the script, so that future
        sessions launch with these settings instead of the hardcoded
        factory defaults."""
        ok, err = self._read_into_params()
        if not ok:
            messagebox.showerror("Invalid parameters", err)
            return
        ok, info = save_user_defaults(self.app.state.params)
        if ok:
            messagebox.showinfo(
                "Defaults saved",
                f"Current parameters were saved as the new default.\n\n"
                f"Stored in:\n  {info}\n\n"
                f"Future sessions will start with these values.")
            self.app.status.configure(
                text=f"Defaults saved to {info}")
        else:
            messagebox.showerror("Could not save defaults",
                                  f"Reason: {info}")

    def _clear_saved_default(self):
        """Delete any saved user defaults file, so the next session
        starts from the hardcoded factory defaults again. Does not
        change the current Parameters-tab values; click 'Restore
        defaults' for that."""
        if not messagebox.askyesno(
                "Clear saved default",
                "Delete the kinetics_defaults.json file?\n\n"
                "Future sessions will start from the hardcoded factory "
                "defaults. The current Parameters-tab values are not "
                "changed by this action."):
            return
        ok, info = clear_user_defaults()
        if ok:
            messagebox.showinfo(
                "Defaults cleared",
                "The saved defaults file has been removed.")
            self.app.status.configure(
                text="Saved defaults cleared")
        else:
            messagebox.showerror("Could not clear defaults",
                                  f"Reason: {info}")

    def _reextract(self):
        ok, err = self._read_into_params()
        if not ok:
            messagebox.showerror("Invalid parameters", err)
            return

        # Warn if any manual adjustments would be lost
        manual_count = sum(
            1 for s in self.app.state.sessions
            for p in s.pairs if p.manually_adjusted)
        if manual_count > 0:
            ok = messagebox.askokcancel(
                "Discard manual adjustments?",
                f"Re-extraction will reset {manual_count} manually-"
                "adjusted fit window(s) to the new auto values. "
                "Concentrations, include flags, [E] and ε will be "
                "preserved.\n\nProceed?")
            if not ok:
                return

        self.app.reextract_all()


class FileTab(ttk.Frame):
    """Per-file tab: per-pair zoom plot on top, slope-pair table below.

    Clicking a plot panel toggles the include flag for that pair. Excluded
    panels are greyed out. Concentration edits in the table re-title the
    corresponding panel without redrawing the trace.
    """

    HEADERS = ["pair", "t_enz (min)", "slope_blank", "slope_enz",
               "slope_net", "R²_enz", "meets", "window (s)", "manual",
               "[S] (µM)", "v (µM/s)", "kobs (1/s)", "include"]

    # Per-panel layout
    _PLOT_NCOLS = 5
    # An edge-zone for drag detection: this fraction of the FIT WINDOW
    # WIDTH on either side (in time) is treated as the draggable edge.
    # Capped at the equivalent of N seconds so very short windows still
    # have a usable click target outside the edge zone.
    _EDGE_FRAC = 0.12
    _EDGE_MIN_SEC = 0.5
    _EDGE_MAX_SEC = 3.0

    def __init__(self, master, app: "App", session: FileSession):
        super().__init__(master)
        self.app = app
        self.session = session
        self._row_widgets: list[dict] = []
        # Per-pair plot bookkeeping (built when entering pair-analysis view)
        self._panel_axes: list = []
        self._panel_titles: list = []
        self._panel_traces: list = []   # tuples (enz_line, blank_line)
        self._panel_fits: list = []     # tuples (enz_fit, blank_fit)
        self._panel_shades: list = []
        self._drag = None
        self._click_suppressed = False
        # Region selection bookkeeping (built when entering region view)
        self._region_canvas = None
        self._region_fig = None
        self._region_ax = None
        self._region_segment_artists: list = []
        self._region_drag = None
        self._region_span_selector = None
        # Lazy-build flag: until `_lazy_build` runs, this tab shows only a
        # placeholder. The build is triggered by the App's notebook tab
        # selection handler the first time the user clicks this tab.
        self._lazy_built = False
        # Placeholder shown until the user clicks this tab. Cheap.
        self._placeholder = ttk.Label(
            self,
            text=f"  {session.name}\n  (click this tab to load)",
            foreground="#666", justify="left")
        self._placeholder.pack(fill="both", expand=True, padx=12, pady=12)

    def lazy_build(self):
        """Build the heavy view content. Called once, the first time the
        user selects this tab. Loads the raw trace, constructs both view
        frames, and shows whichever view the session state points at."""
        if self._lazy_built:
            return
        self._lazy_built = True
        # Drop placeholder
        try:
            self._placeholder.destroy()
        except Exception:
            pass

        # Load the raw trace if not already cached on the session
        if getattr(self.session, "_raw_trace", None) is None:
            try:
                df = kp.load_cary_csv(self.session.path)
                self.session._raw_trace = (
                    df["time_min"].to_numpy(), df["abs"].to_numpy())
            except Exception as exc:
                self.session._raw_trace = None
                self.session.error = f"Could not load trace: {exc}"

        # ---------------- common header / view toggle ----------------
        head = ttk.Frame(self)
        head.pack(fill="x", padx=8, pady=(8, 2))
        ttk.Label(head, text=self.session.name,
                  font=("TkDefaultFont", 11, "bold")).pack(side="left")
        self._n_pairs_lbl = ttk.Label(head, text="")
        self._n_pairs_lbl.pack(side="left", padx=(8, 0))
        if self.session.error:
            ttk.Label(head, text="  !!! " + self.session.error,
                      foreground="red").pack(side="left")

        toggle_bar = ttk.Frame(self)
        toggle_bar.pack(fill="x", padx=8, pady=(0, 4))
        self._view_var = tk.StringVar(
            value=("pair" if self.session.pairs_confirmed else "region"))
        ttk.Radiobutton(toggle_bar, text="1. Region selection",
                        value="region", variable=self._view_var,
                        command=self._switch_view).pack(side="left")
        ttk.Radiobutton(toggle_bar, text="2. Pair analysis",
                        value="pair", variable=self._view_var,
                        command=self._switch_view).pack(side="left",
                                                         padx=(12, 0))
        self._confirm_state_lbl = ttk.Label(toggle_bar, text="",
                                            foreground="#666")
        self._confirm_state_lbl.pack(side="left", padx=(20, 0))

        self._region_frame = ttk.Frame(self)
        self._pair_frame = ttk.Frame(self)

        self._refresh_confirm_state_label()
        self._update_pair_count_label()
        self._switch_view()

    # ------------------------------------------------------------------
    # View switching
    # ------------------------------------------------------------------

    def _switch_view(self):
        """Show the view selected by the radio. The non-selected frame is
        forgotten (not destroyed) so we can reuse it."""
        which = self._view_var.get()
        if which == "pair" and not self.session.pairs_confirmed:
            # Block entry into pair analysis until confirmed
            self._view_var.set("region")
            messagebox.showinfo(
                "Confirm pairs first",
                "Please review and confirm the regions in step 1 before "
                "advancing to pair analysis.")
            which = "region"

        if which == "region":
            self._pair_frame.pack_forget()
            if not self._region_frame.winfo_children():
                self._build_region_selection_view()
            self._region_frame.pack(fill="both", expand=True)
            self._refresh_region_view()
        else:
            self._region_frame.pack_forget()
            # Always rebuild the pair-analysis view: pairs may have
            # changed since last time we showed it.
            for w in self._pair_frame.winfo_children():
                w.destroy()
            self._reset_pair_view_state()
            self._build_pair_analysis_view()
            self._pair_frame.pack(fill="both", expand=True)

    def _reset_pair_view_state(self):
        self._row_widgets = []
        self._panel_axes = []
        self._panel_titles = []
        self._panel_traces = []
        self._panel_fits = []
        self._panel_shades = []
        self._drag = None
        self._click_suppressed = False

    def _refresh_confirm_state_label(self):
        if self.session.pairs_confirmed:
            self._confirm_state_lbl.configure(
                text="✓ pairs confirmed", foreground="#1a7a1a")
        else:
            self._confirm_state_lbl.configure(
                text="(pairs not yet confirmed)", foreground="#a04000")

    def _update_pair_count_label(self):
        n_seg_b = sum(1 for s in self.session.segments if s.label == "blank")
        n_seg_e = sum(1 for s in self.session.segments if s.label == "enzyme")
        n_pairs = len(self.session.pairs)
        self._n_pairs_lbl.configure(
            text=f"   ({n_seg_b} blanks, {n_seg_e} enzymes → {n_pairs} pairs)")

    # ==================================================================
    # STEP 2: pair analysis view (the existing UI)
    # ==================================================================

    def _build_pair_analysis_view(self):
        """The original FileTab content — per-pair plot + table."""
        parent = self._pair_frame
        session = self.session

        # Live result strip
        self._result_lbl = ttk.Label(parent, text="", anchor="w",
                                     font=("TkDefaultFont", 10, "bold"))
        self._result_lbl.pack(fill="x", padx=8, pady=2)

        # Quick actions
        bar = ttk.Frame(parent)
        bar.pack(fill="x", padx=8, pady=2)
        ttk.Button(bar, text="Include all",
                   command=lambda: self._set_all_included(True)
                   ).pack(side="left", padx=2)
        ttk.Button(bar, text="Exclude all",
                   command=lambda: self._set_all_included(False)
                   ).pack(side="left", padx=2)
        ttk.Button(bar, text="Reset concentrations",
                   command=self._reset_concentrations
                   ).pack(side="left", padx=8)
        ttk.Label(bar,
                  text=("(click panel = toggle inclusion;  drag panel "
                        "edge = adjust fit window;  right-click = reset "
                        "to auto)"),
                  foreground="grey"
                  ).pack(side="left", padx=12)

        # Plot + table
        self._build_plot(parent=parent)
        scroll = ScrollableFrame(parent)
        scroll.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        tbl = scroll.inner
        for j, h in enumerate(self.HEADERS):
            ttk.Label(tbl, text=h, font=("TkDefaultFont", 9, "bold")
                      ).grid(row=0, column=j, padx=4, pady=2, sticky="w")
        for i, p in enumerate(session.pairs, start=1):
            self._make_row(tbl, i, p)
        self.update_row_displays()
        self.update_result_label()

    # ==================================================================
    # STEP 1: region selection view
    # ==================================================================

    # Edge-zone width for region drag (fraction of segment width)
    _REGION_EDGE_FRAC = 0.15
    _REGION_EDGE_MIN_SEC = 0.5
    _REGION_EDGE_MAX_SEC = 5.0

    def _build_region_selection_view(self):
        """Big trace, lasso-add, drag-edges, click-flip, right-click delete."""
        parent = self._region_frame

        info = ttk.Label(parent, text=(
            "Step 1: review the auto-detected segments. Drag a segment's "
            "edge to retime it. Right-click a segment to delete. "
            "Lasso a new range on the plot, then click 'Add as blank' or "
            "'Add as enzyme' to insert a new segment."),
            wraplength=1200, justify="left", foreground="#444")
        info.pack(fill="x", padx=8, pady=(2, 6))

        # Action bar
        bar = ttk.Frame(parent)
        bar.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Button(bar, text="Re-run auto-detect",
                   command=self._region_redetect).pack(side="left")
        ttk.Button(bar, text="Clear all segments",
                   command=self._region_clear_all).pack(side="left",
                                                         padx=(8, 0))
        ttk.Label(bar, text="    Pending lasso →").pack(side="left",
                                                          padx=(20, 4))
        self._region_pending_lbl = ttk.Label(bar, text="(none)",
                                             foreground="#888")
        self._region_pending_lbl.pack(side="left")
        self._region_add_blank_btn = ttk.Button(
            bar, text="Add as blank", state="disabled",
            command=lambda: self._region_commit_lasso("blank"))
        self._region_add_blank_btn.pack(side="left", padx=(8, 2))
        self._region_add_enzyme_btn = ttk.Button(
            bar, text="Add as enzyme", state="disabled",
            command=lambda: self._region_commit_lasso("enzyme"))
        self._region_add_enzyme_btn.pack(side="left", padx=2)
        self._region_cancel_btn = ttk.Button(
            bar, text="Cancel lasso", state="disabled",
            command=self._region_cancel_lasso)
        self._region_cancel_btn.pack(side="left", padx=2)

        ttk.Button(bar, text="Confirm pairs →",
                   command=self._region_confirm,
                   ).pack(side="right")

        # Big trace
        if self.session._raw_trace is None:
            ttk.Label(parent, text="(raw trace unavailable)",
                      foreground="red").pack(fill="x", padx=8, pady=20)
            return

        t, a = self.session._raw_trace
        fig = Figure(figsize=(14, 5.0), dpi=100)
        ax = fig.add_subplot(111)
        ax.plot(t, a, color="#888", linewidth=0.6, alpha=0.85)
        ax.set_xlabel("Time (min)")
        ax.set_ylabel("Absorbance")
        ax.grid(alpha=0.3)
        # Reasonable y-bounds based on trace
        a_min, a_max = float(np.min(a)), float(np.max(a))
        margin = 0.04 * (a_max - a_min if a_max > a_min else 1.0)
        ax.set_ylim(max(-0.05, a_min - margin), min(0.7, a_max + margin))
        ax.set_xlim(float(t[0]), float(t[-1]))
        self._region_fig = fig
        self._region_ax = ax
        self._region_canvas = FigureCanvasTkAgg(fig, master=parent)
        widget = self._region_canvas.get_tk_widget()
        widget.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk
        toolbar = NavigationToolbar2Tk(self._region_canvas, parent)
        toolbar.update()
        toolbar.pack(side="top", fill="x", padx=8)

        # SpanSelector for lasso-add. NOTE: useblit=False is required —
        # with useblit=True, the selector's internal artists get
        # invalidated when we mutate `axvspan`/`text` artists in
        # `_region_render_segments`, leaving the selector inert after
        # the first commit.
        from matplotlib.widgets import SpanSelector
        self._region_pending_span: Optional[tuple[float, float]] = None
        self._region_span_selector = SpanSelector(
            ax, self._on_region_lasso, "horizontal",
            useblit=False, props=dict(alpha=0.3, facecolor="#ffd54f"),
            interactive=False, ignore_event_outside=False,
            minspan=0.05,
        )

        # Mouse handlers for edge drag / right-click delete on segments
        fig.canvas.mpl_connect("button_press_event",
                               self._on_region_mouse_press)
        fig.canvas.mpl_connect("motion_notify_event",
                               self._on_region_mouse_motion)
        fig.canvas.mpl_connect("button_release_event",
                               self._on_region_mouse_release)

        self._region_render_segments()

    def _refresh_region_view(self):
        """Called when re-entering the region tab — refresh visuals."""
        if self._region_ax is None:
            return
        self._region_render_segments()

    def _region_render_segments(self):
        """Redraw the per-segment shaded boxes + labels.

        Naming convention: blanks and enzymes that form a pair share a
        number. Pair number N is assigned in time order across enzymes;
        the blank that serves pair N (the most recently-preceding blank
        of enzyme N) gets the same number. A blank that serves multiple
        enzymes (replicates) keeps the number of the first enzyme it
        served.
        """
        if self._region_ax is None:
            return
        # Clear previous artists
        for span, label in self._region_segment_artists:
            try:
                span.remove()
            except Exception:
                pass
            try:
                label.remove()
            except Exception:
                pass
        self._region_segment_artists = []

        if self.session._raw_trace is None:
            return
        t, _ = self.session._raw_trace
        segs_sorted = sorted(self.session.segments, key=lambda sp: sp.s_idx)

        # Walk in time order, assigning pair numbers
        pair_no = 0
        last_blank_idx: Optional[int] = None
        # blank_pair_num[i] = pair number to display for segment at sorted index i
        seg_pair_nums: list[Optional[int]] = [None] * len(segs_sorted)
        for i, sp in enumerate(segs_sorted):
            if sp.label == "blank":
                last_blank_idx = i
                continue
            if sp.label != "enzyme":
                continue
            pair_no += 1
            seg_pair_nums[i] = pair_no
            if last_blank_idx is not None and seg_pair_nums[last_blank_idx] is None:
                seg_pair_nums[last_blank_idx] = pair_no

        for i, sp in enumerate(segs_sorted):
            if sp.s_idx < 0 or sp.e_idx > len(t) or sp.e_idx <= sp.s_idx:
                continue
            t_lo = float(t[sp.s_idx])
            t_hi = float(t[min(sp.e_idx, len(t) - 1)])
            color = "tab:blue" if sp.label == "blank" else "tab:red"
            span = self._region_ax.axvspan(t_lo, t_hi, color=color,
                                           alpha=0.18, picker=False)
            mid = 0.5 * (t_lo + t_hi)
            ymin, ymax = self._region_ax.get_ylim()
            num = seg_pair_nums[i]
            if num is None:
                # Orphan blank with no following enzyme yet — flag as '?'
                tag = f"{sp.label[0].upper()}?"
            else:
                tag = f"{sp.label[0].upper()}{num}"
            label = self._region_ax.text(
                mid, ymax - 0.04 * (ymax - ymin),
                tag,
                ha="center", va="top", fontsize=8,
                bbox=dict(boxstyle="round,pad=0.2",
                          facecolor=color, alpha=0.6,
                          edgecolor="none"),
                color="white")
            self._region_segment_artists.append((span, label))
        self._region_canvas.draw_idle()
        self._update_pair_count_label()

    # -- lasso (SpanSelector) flow --

    def _on_region_lasso(self, t_lo: float, t_hi: float):
        """SpanSelector callback when the user finishes a click-drag."""
        self._region_pending_span = (float(t_lo), float(t_hi))
        self._region_pending_lbl.configure(
            text=f"{t_lo:.2f}–{t_hi:.2f} min",
            foreground="#000")
        self._region_add_blank_btn.configure(state="normal")
        self._region_add_enzyme_btn.configure(state="normal")
        self._region_cancel_btn.configure(state="normal")

    def _region_commit_lasso(self, label: str):
        if self._region_pending_span is None or self.session._raw_trace is None:
            return
        t, _ = self.session._raw_trace
        t_lo, t_hi = self._region_pending_span
        s_idx = int(np.searchsorted(t, t_lo))
        e_idx = int(np.searchsorted(t, t_hi))
        if e_idx <= s_idx + 5:
            messagebox.showwarning(
                "Range too small",
                "The lasso range is too short to be a real measurement "
                "(needs at least 5 samples).")
            return
        self.session.segments.append(SegmentSpec(
            s_idx=s_idx, e_idx=e_idx, label=label))
        # Mark unconfirmed (we changed something)
        if self.session.pairs_confirmed:
            self.session.pairs_confirmed = False
            self._refresh_confirm_state_label()
        self._region_cancel_lasso()
        self._region_render_segments()
        # Recreate the SpanSelector. Some matplotlib versions leave the
        # selector in a dead state after the first commit; rebuilding
        # it guarantees it remains responsive for further selections.
        self._region_recreate_span_selector()

    def _region_recreate_span_selector(self):
        """Tear down and rebuild the SpanSelector."""
        if self._region_ax is None:
            return
        from matplotlib.widgets import SpanSelector
        if self._region_span_selector is not None:
            try:
                self._region_span_selector.set_active(False)
                self._region_span_selector.disconnect_events()
            except Exception:
                pass
        self._region_span_selector = SpanSelector(
            self._region_ax, self._on_region_lasso, "horizontal",
            useblit=False, props=dict(alpha=0.3, facecolor="#ffd54f"),
            interactive=False, ignore_event_outside=False,
            minspan=0.05,
        )

    def _region_cancel_lasso(self):
        self._region_pending_span = None
        self._region_pending_lbl.configure(text="(none)", foreground="#888")
        self._region_add_blank_btn.configure(state="disabled")
        self._region_add_enzyme_btn.configure(state="disabled")
        self._region_cancel_btn.configure(state="disabled")
        # Make sure the selector is alive for the next attempt
        if self._region_ax is not None and self._region_span_selector is not None:
            try:
                self._region_span_selector.set_active(True)
            except Exception:
                pass

    # -- segment edge drag / right-click delete --

    def _region_hit_test(self, x_min: float
                         ) -> tuple[Optional[int], Optional[str]]:
        """Return (segment_index, edge) where edge is 'left'/'right'/None.
        index is into `self.session.segments` (the unsorted list)."""
        if self.session._raw_trace is None:
            return None, None
        t, _ = self.session._raw_trace
        best: tuple[Optional[int], Optional[str], float] = (None, None, 1e9)
        for k, sp in enumerate(self.session.segments):
            if sp.s_idx < 0 or sp.e_idx > len(t):
                continue
            t_lo = float(t[sp.s_idx])
            t_hi = float(t[min(sp.e_idx, len(t) - 1)])
            width = t_hi - t_lo
            zone = max(self._REGION_EDGE_MIN_SEC / 60.0,
                       min(self._REGION_EDGE_MAX_SEC / 60.0,
                           self._REGION_EDGE_FRAC * width))
            d_l = abs(x_min - t_lo)
            d_r = abs(x_min - t_hi)
            if d_l < zone and d_l < best[2]:
                best = (k, "left", d_l)
            if d_r < zone and d_r < best[2]:
                best = (k, "right", d_r)
        return best[0], best[1]

    def _region_segment_at(self, x_min: float) -> Optional[int]:
        """Return index of the segment whose interior contains x_min."""
        if self.session._raw_trace is None:
            return None
        t, _ = self.session._raw_trace
        for k, sp in enumerate(self.session.segments):
            if sp.s_idx < 0 or sp.e_idx > len(t):
                continue
            if t[sp.s_idx] <= x_min <= t[min(sp.e_idx, len(t) - 1)]:
                return k
        return None

    def _on_region_mouse_press(self, event):
        if event.inaxes is not self._region_ax:
            return
        if event.xdata is None:
            return

        # Right-click: delete the segment under the cursor.
        if event.button == 3:
            k = self._region_segment_at(event.xdata)
            if k is not None:
                if messagebox.askyesno(
                    "Delete segment?",
                    f"Delete this {self.session.segments[k].label} segment?"):
                    del self.session.segments[k]
                    if self.session.pairs_confirmed:
                        self.session.pairs_confirmed = False
                        self._refresh_confirm_state_label()
                    self._region_render_segments()
            return

        if event.button != 1:
            return

        # Edge drag detection
        k, edge = self._region_hit_test(event.xdata)
        if k is None:
            return
        # Disable the SpanSelector while dragging an edge so it doesn't
        # eat the drag.
        if self._region_span_selector is not None:
            self._region_span_selector.set_active(False)
        self._region_drag = {
            "k": k, "edge": edge,
            "press_x": event.xdata,
        }

    def _on_region_mouse_motion(self, event):
        if self._region_drag is None:
            return
        if event.inaxes is not self._region_ax or event.xdata is None:
            return
        if self.session._raw_trace is None:
            return
        t, _ = self.session._raw_trace
        k = self._region_drag["k"]
        edge = self._region_drag["edge"]
        sp = self.session.segments[k]
        new_idx = int(np.searchsorted(t, event.xdata))
        new_idx = max(0, min(len(t), new_idx))
        if edge == "left":
            new_idx = max(0, min(new_idx, sp.e_idx - 5))
            sp.s_idx = new_idx
        else:
            new_idx = max(sp.s_idx + 5, min(len(t), new_idx))
            sp.e_idx = new_idx
        self._region_render_segments()

    def _on_region_mouse_release(self, event):
        if self._region_drag is None:
            return
        self._region_drag = None
        if self._region_span_selector is not None:
            self._region_span_selector.set_active(True)
        if self.session.pairs_confirmed:
            self.session.pairs_confirmed = False
            self._refresh_confirm_state_label()
        self._region_render_segments()

    # -- top-level region actions --

    def _region_redetect(self):
        if self.session.pairs_confirmed and self.session.segments:
            ok = messagebox.askyesno(
                "Discard manual edits?",
                "Re-running auto-detect will replace all current segments "
                "with the algorithm's output. Continue?")
            if not ok:
                return
        new_segs, err = auto_detect_segments(
            self.session.path, params=self.app.state.params)
        if err and not new_segs:
            messagebox.showerror("Auto-detect failed", err)
            return
        self.session.segments = new_segs
        self.session.pairs_confirmed = False
        self._refresh_confirm_state_label()
        self._region_render_segments()

    def _region_clear_all(self):
        if not self.session.segments:
            return
        if not messagebox.askyesno(
            "Clear all segments?",
            "Remove all segments from this variant? You can re-run auto-"
            "detect or lasso new ones."):
            return
        self.session.segments = []
        if self.session.pairs_confirmed:
            self.session.pairs_confirmed = False
            self._refresh_confirm_state_label()
        self._region_render_segments()

    def _region_confirm(self):
        """Build pairs from segments, set pairs_confirmed, persist sidecar,
        and switch to the pair-analysis view."""
        if not self.session.segments:
            messagebox.showwarning(
                "No segments",
                "There are no segments to confirm. Lasso some regions or "
                "run auto-detect first.")
            return
        # Sort segments by time
        self.session.segments.sort(key=lambda sp: sp.s_idx)
        pairs, err = build_pairs_from_segments(
            self.session.path, self.session.segments,
            params=self.app.state.params)
        if err:
            # Non-fatal; show as a status message but continue
            self.session.error = err
        self.session.pairs = pairs
        self.session.pairs_confirmed = True
        recompute_session(self.session, self.app.state.params)
        self._refresh_confirm_state_label()
        self._update_pair_count_label()
        self.app.on_session_confirmed(self.session)
        # Switch radio + view
        self._view_var.set("pair")
        self._switch_view()

    # ==================================================================
    # STEP 2: pair-analysis view methods (unchanged from before)
    # ==================================================================

    def _build_plot(self, parent=None):
        """Build the per-pair zoom subplots once. Each panel shows the
        enzyme trace + fit (red/black) and the blank trace + fit
        (steelblue/dimgrey, baseline-shifted to start at the enzyme
        onset absorbance so the slopes can be compared visually).
        Per-panel handles are stored so toggling include or editing a
        concentration only mutates the affected panel.
        """
        if parent is None:
            parent = self
        if not self.session.pairs:
            return

        # Need the raw trace from the source CSV to plot the segments.
        if getattr(self.session, "_raw_trace", None) is None:
            try:
                df = kp.load_cary_csv(self.session.path)
                self.session._raw_trace = (
                    df["time_min"].to_numpy(), df["abs"].to_numpy())
            except Exception as exc:
                ttk.Label(parent,
                          text=f"(could not load raw trace for plot: {exc})",
                          foreground="red"
                          ).pack(fill="x", padx=8, pady=4)
                self.session._raw_trace = None
                return

        t, a = self.session._raw_trace

        # Use each pair's own stashed segment indices. This honours the
        # user's confirmed regions exactly.
        enz_segs = [(p._enzyme_seg or (0, 0)) for p in self.session.pairs]
        blank_segs = [(p._blank_seg or (0, 0)) for p in self.session.pairs]

        n_pairs = len(self.session.pairs)
        ncols = self._PLOT_NCOLS
        nrows = int(np.ceil(n_pairs / ncols))

        fig = Figure(figsize=(2.4 * ncols, 1.9 * nrows), dpi=100)
        for idx, p in enumerate(self.session.pairs):
            ax = fig.add_subplot(nrows, ncols, idx + 1)
            self._panel_axes.append(ax)

            # Enzyme segment slice
            if idx < len(enz_segs):
                s_e, e_e = enz_segs[idx]
                xs_enz = t[s_e:e_e]
                ys_enz = a[s_e:e_e]
            else:
                xs_enz = np.array([p.t_enzyme_start])
                ys_enz = np.array([0.0])

            # Plot enzyme trace + fit (with the y-anchor from the trace
            # value at the fit-window start, see comment further down).
            (line_enz_trace,) = ax.plot(
                xs_enz, ys_enz, color="tab:red", linewidth=0.9, alpha=0.85,
                label="enzyme")
            t_fit_start = p.t_enzyme_start
            t_fit_end = p.t_enzyme_start + p.window_enzyme_s / 60.0
            anchor_idx = int(np.searchsorted(xs_enz, t_fit_start))
            anchor_idx = min(max(anchor_idx, 0), len(ys_enz) - 1)
            a_anchor = ys_enz[anchor_idx] if len(ys_enz) else 0.0
            xs_fit = np.array([t_fit_start, t_fit_end])
            ys_fit = a_anchor + p.slope_enzyme * (xs_fit - t_fit_start)
            (line_enz_fit,) = ax.plot(
                xs_fit, ys_fit, color="black", linewidth=1.6,
                label="enzyme fit")
            shade = ax.axvspan(t_fit_start, t_fit_end,
                               color="black", alpha=0.08)

            # Blank segment slice. Shift it vertically so it starts at
            # the enzyme onset absorbance: this lets the user compare
            # the SLOPE of the blank against the slope of the enzyme on
            # one axis without the panel y-range being dominated by the
            # absolute offset between the two segments.
            line_blank_trace = None
            line_blank_fit = None
            if idx < len(blank_segs):
                s_b, e_b = blank_segs[idx]
                xs_bl = t[s_b:e_b]
                ys_bl_raw = a[s_b:e_b]
                if len(ys_bl_raw):
                    # Choose anchor: blank is shifted so its FIRST point
                    # aligns vertically with the enzyme onset.
                    bl_offset = ys_enz[0] - ys_bl_raw[0] if len(ys_enz) else 0.0
                    ys_bl = ys_bl_raw + bl_offset
                    (line_blank_trace,) = ax.plot(
                        xs_bl, ys_bl, color="tab:blue",
                        linewidth=0.9, alpha=0.7, label="blank (shifted)")
                    # Blank fit: drawn over the EFFECTIVE fit window
                    # (which excludes any post-spike transient skipped
                    # by `_find_blank_settled_start`). The y values are
                    # in the same shifted frame as the blank trace.
                    if p._blank_fit_seg is not None:
                        sb_fit, eb_fit = p._blank_fit_seg
                        x_fl = t[sb_fit:eb_fit]
                        if len(x_fl) >= 2:
                            xs_blf = np.array([x_fl[0], x_fl[-1]])
                        else:
                            xs_blf = np.array([xs_bl[0], xs_bl[-1]])
                    else:
                        xs_blf = np.array([xs_bl[0], xs_bl[-1]])
                    # Anchor the line on the actual trace value at fit
                    # start (in the shifted frame). This handles the
                    # vertical shift correctly even when there's a head
                    # transient that we skipped.
                    anchor_idx = int(np.searchsorted(xs_bl, xs_blf[0]))
                    anchor_idx = min(max(anchor_idx, 0), len(ys_bl) - 1)
                    a_anchor_bl = ys_bl[anchor_idx]
                    ys_blf = a_anchor_bl + p.slope_blank * (xs_blf - xs_blf[0])
                    (line_blank_fit,) = ax.plot(
                        xs_blf, ys_blf, color="dimgrey",
                        linewidth=1.4, linestyle="--",
                        label="blank fit")

            self._panel_traces.append((line_enz_trace, line_blank_trace))
            self._panel_fits.append((line_enz_fit, line_blank_fit))
            self._panel_shades.append(shade)

            ax.tick_params(labelsize=7)
            ax.grid(alpha=0.3)
            title = ax.set_title(self._panel_title_for(p), fontsize=8)
            self._panel_titles.append(title)
            ax.set_picker(True)

        # One legend at the figure level, top-right corner of figure.
        # We collect handles from the first panel that has all four lines.
        legend_handles = None
        legend_labels = None
        for ax in fig.axes:
            handles, labs = ax.get_legend_handles_labels()
            if len(handles) >= 4:
                legend_handles, legend_labels = handles, labs
                break
        if legend_handles is not None:
            fig.legend(legend_handles, legend_labels,
                       loc="upper right", fontsize=7,
                       bbox_to_anchor=(0.995, 0.995),
                       framealpha=0.9, ncol=4)

        fig.tight_layout(rect=(0, 0, 1, 0.97))
        self._fig = fig
        self._canvas = FigureCanvasTkAgg(fig, master=parent)
        widget = self._canvas.get_tk_widget()
        widget.configure(height=1.9 * nrows * 100)
        widget.pack(fill="x", padx=8, pady=(2, 4))

        self._fig.canvas.mpl_connect(
            "button_press_event", self._on_mouse_press)
        self._fig.canvas.mpl_connect(
            "motion_notify_event", self._on_mouse_motion)
        self._fig.canvas.mpl_connect(
            "button_release_event", self._on_mouse_release)
        self._refresh_panel_styles()

    def _panel_title_for(self, p: PairRow) -> str:
        s = self.session
        v = velocity_uM_per_s(p.slope_net, s.epsilon_M1cm1)
        k = kobs_per_s(v, s.enzyme_conc_uM)
        conc = (f"[S]={p.concentration_uM:g}"
                if np.isfinite(p.concentration_uM) else "[S]=?")
        meets = "" if p.meets_r2 else " ⚠"
        kobs_part = (f"  kobs={k:.3g}/s"
                     if np.isfinite(k) else "")
        return f"#{p.pair} {conc}µM{meets}{kobs_part}"

    def _refresh_panel_styles(self):
        """Apply 'included' vs 'excluded' visual styling to every panel.
        Manually-adjusted panels get a blue border."""
        if not self._panel_axes:
            return
        for p, ax, trace_pair, fit_pair, shade in zip(
                self.session.pairs, self._panel_axes,
                self._panel_traces, self._panel_fits, self._panel_shades):
            line_enz_trace, line_blank_trace = trace_pair
            line_enz_fit, line_blank_fit = fit_pair
            if p.include:
                if line_enz_trace is not None:
                    line_enz_trace.set_color("tab:red")
                    line_enz_trace.set_alpha(0.85)
                if line_blank_trace is not None:
                    line_blank_trace.set_color("tab:blue")
                    line_blank_trace.set_alpha(0.7)
                if line_enz_fit is not None:
                    line_enz_fit.set_color("black")
                    line_enz_fit.set_linewidth(1.6)
                if line_blank_fit is not None:
                    line_blank_fit.set_color("dimgrey")
                    line_blank_fit.set_linewidth(1.4)
                shade.set_alpha(0.08)
                ax.set_facecolor("white")
            else:
                if line_enz_trace is not None:
                    line_enz_trace.set_color("#cccccc")
                    line_enz_trace.set_alpha(0.5)
                if line_blank_trace is not None:
                    line_blank_trace.set_color("#cccccc")
                    line_blank_trace.set_alpha(0.5)
                if line_enz_fit is not None:
                    line_enz_fit.set_color("#bbbbbb")
                    line_enz_fit.set_linewidth(1.0)
                if line_blank_fit is not None:
                    line_blank_fit.set_color("#bbbbbb")
                    line_blank_fit.set_linewidth(1.0)
                shade.set_alpha(0.0)
                ax.set_facecolor("#f4f4f4")
            # Spine colour communicates manual-adjustment status
            if p.manually_adjusted:
                spine_color = "tab:blue"
                spine_lw = 1.6
            elif p.include:
                spine_color = "#444"
                spine_lw = 0.8
            else:
                spine_color = "#aaaaaa"
                spine_lw = 0.8
            for spine in ax.spines.values():
                spine.set_color(spine_color)
                spine.set_linewidth(spine_lw)
        # Synchronous draw — matplotlib's draw_idle defers redraws until
        # the Tk event loop is idle, which on some platforms means "until
        # the user clicks something else first". For one-off edits we
        # want the change visible immediately.
        debug_draw("_refresh_panel_styles", "before draw()")
        self._canvas.draw()
        debug_draw("_refresh_panel_styles", "after draw()")

    def _refresh_panel_titles(self):
        for p, title in zip(self.session.pairs, self._panel_titles):
            title.set_text(self._panel_title_for(p))
        debug_draw("_refresh_panel_titles", "before draw()")
        self._canvas.draw()
        debug_draw("_refresh_panel_titles", "after draw()")

    # ------------------------------------------------------------------
    # Mouse interaction: edge-drag to adjust fit window; click in
    # interior to toggle inclusion; right-click to reset to auto.
    # ------------------------------------------------------------------

    def _edge_zone_width(self, t_start: float, t_end: float) -> float:
        """Width of the edge-detection zone in minutes."""
        win_min = max(0.0, t_end - t_start)
        zone = self._EDGE_FRAC * win_min
        zone = min(zone, self._EDGE_MAX_SEC / 60.0)
        zone = max(zone, self._EDGE_MIN_SEC / 60.0)
        return zone

    def _hit_test_edge(self, ax_idx: int, x_min: float
                       ) -> Optional[str]:
        """Return 'left', 'right', or None depending on whether `x_min`
        is within the edge zone of the panel's fit window."""
        if ax_idx >= len(self.session.pairs):
            return None
        p = self.session.pairs[ax_idx]
        t_start = p.t_enzyme_start
        t_end = p.t_enzyme_start + p.window_enzyme_s / 60.0
        zone = self._edge_zone_width(t_start, t_end)
        if abs(x_min - t_start) <= zone:
            return "left"
        if abs(x_min - t_end) <= zone:
            return "right"
        return None

    def _ax_index(self, ax) -> Optional[int]:
        try:
            return self._panel_axes.index(ax)
        except ValueError:
            return None

    def _on_mouse_press(self, event):
        if event.inaxes is None:
            return
        idx = self._ax_index(event.inaxes)
        if idx is None:
            return

        # Double-click (left): open the pair in a zoomable popout window
        if event.button == 1 and getattr(event, "dblclick", False):
            self._open_panel_popout(idx)
            self._click_suppressed = True
            return

        # Right click: reset this pair to auto (or no-op if already auto)
        if event.button == 3:
            p = self.session.pairs[idx]
            if p.manually_adjusted:
                revert_pair_to_auto(p)
                self._update_panel_for_pair(idx)
                recompute_session(self.session, self.app.state.params)
                self.update_row_displays()
                self._refresh_panel_styles()
                self._refresh_panel_titles()
                self.update_result_label()
                self.app.recomputed_one(self.session, refresh_self=False)
                self._click_suppressed = True
            return

        if event.button != 1:
            return

        # Edge-drag detection
        if event.xdata is None:
            return
        edge = self._hit_test_edge(idx, event.xdata)
        if edge is None:
            return  # falls through to a possible click-toggle on release

        p = self.session.pairs[idx]
        t_start = p.t_enzyme_start
        t_end = p.t_enzyme_start + p.window_enzyme_s / 60.0
        self._drag = {
            "idx": idx,
            "edge": edge,
            "t_other": t_end if edge == "left" else t_start,
            "did_move": False,
            "press_x": event.xdata,
        }

    def _on_mouse_motion(self, event):
        if self._drag is None:
            return
        if event.inaxes is None:
            return
        idx = self._drag["idx"]
        if self._ax_index(event.inaxes) != idx:
            return
        if event.xdata is None:
            return
        self._drag["did_move"] = True
        # Update the shaded region live without refitting
        new_x = event.xdata
        t_other = self._drag["t_other"]
        t_lo = min(new_x, t_other)
        t_hi = max(new_x, t_other)
        # Clamp to the enzyme segment bounds
        p = self.session.pairs[idx]
        if p._enzyme_seg is not None and self.session._raw_trace is not None:
            t_full, _ = self.session._raw_trace
            s_e, e_e = p._enzyme_seg
            t_lo = max(t_lo, float(t_full[s_e]))
            t_hi = min(t_hi, float(t_full[e_e - 1]))
        self._set_shade_bounds(idx, t_lo, t_hi)
        self._canvas.draw_idle()

    def _on_mouse_release(self, event):
        # Resolve any in-progress drag
        if self._drag is not None:
            idx = self._drag["idx"]
            did_move = self._drag["did_move"]
            edge = self._drag["edge"]
            t_other = self._drag["t_other"]
            self._drag = None

            if did_move and event.xdata is not None:
                debug_draw("_on_mouse_release", f"drag end idx={idx}")
                t_lo = min(event.xdata, t_other)
                t_hi = max(event.xdata, t_other)
                p = self.session.pairs[idx]
                if self.session._raw_trace is not None:
                    t_full, a_full = self.session._raw_trace
                    ok = refit_pair_window(
                        p, t_full, a_full, t_lo, t_hi,
                        r2_thresh=self.app.state.params.r2_thresh)
                    if ok:
                        debug_draw("_on_mouse_release",
                                    f"refit ok, updating panel idx={idx}")
                        self._update_panel_for_pair(idx)
                        recompute_session(self.session, self.app.state.params)
                        self.update_row_displays()
                        self._refresh_panel_styles()
                        self._refresh_panel_titles()
                        self.update_result_label()
                        self.app.recomputed_one(
                            self.session, refresh_self=False)
                self._click_suppressed = True
                return
            # else: fall through to click-toggle handling below

        # Plain click (no drag): toggle inclusion if click landed on a panel
        if self._click_suppressed:
            self._click_suppressed = False
            return
        if event.inaxes is None or event.button != 1:
            return
        idx = self._ax_index(event.inaxes)
        if idx is None:
            return
        # Don't treat clicks on the edge zones as toggles either
        if event.xdata is not None:
            if self._hit_test_edge(idx, event.xdata) is not None:
                return
        p = self.session.pairs[idx]
        p.include = not p.include
        if idx < len(self._row_widgets):
            self._row_widgets[idx]["inc_var"].set(p.include)
        recompute_session(self.session, self.app.state.params)
        self.update_row_displays()
        self._refresh_panel_styles()
        self._refresh_panel_titles()
        self.update_result_label()
        self.app.recomputed_one(self.session, refresh_self=False)

    def _set_shade_bounds(self, idx: int, t_lo: float, t_hi: float) -> None:
        """Update the axvspan polygon in place (live during drag)."""
        shade = self._panel_shades[idx]
        # axvspan returns a Polygon with 4 vertices forming the rectangle.
        # Rebuild the xy data preserving the y-range from the existing
        # polygon (which is in axes coordinates -- 0 to 1 -- by default).
        xy = shade.get_xy()
        # xy columns: x, y; rows are the polygon vertices (typically 5
        # for a closed rectangle). Mutate the x values.
        new_xy = xy.copy()
        # Pattern: (t_lo, y0), (t_lo, y1), (t_hi, y1), (t_hi, y0), (t_lo, y0)
        new_xy[:, 0] = [t_lo, t_lo, t_hi, t_hi, t_lo][:len(new_xy)]
        shade.set_xy(new_xy)

    def _update_panel_for_pair(self, idx: int) -> None:
        """After a manual refit (or revert) for pair index `idx`, redraw
        the enzyme fit line and the shaded fit window for that panel."""
        if idx >= len(self.session.pairs):
            return
        p = self.session.pairs[idx]
        if self.session._raw_trace is None:
            return
        t_full, a_full = self.session._raw_trace
        if p._enzyme_seg is None:
            return
        s_e, e_e = p._enzyme_seg
        xs = t_full[s_e:e_e]
        ys = a_full[s_e:e_e]
        t_fs = p.t_enzyme_start
        t_fe = p.t_enzyme_start + p.window_enzyme_s / 60.0
        # Reanchor the fit line on the actual trace value at fit start
        anchor_idx = int(np.searchsorted(xs, t_fs))
        anchor_idx = min(max(anchor_idx, 0), len(ys) - 1)
        a_anchor = ys[anchor_idx] if len(ys) else 0.0
        fit_line = self._panel_fits[idx][0]
        if fit_line is not None:
            fit_line.set_data(
                [t_fs, t_fe],
                [a_anchor, a_anchor + p.slope_enzyme * (t_fe - t_fs)],
            )
        self._set_shade_bounds(idx, t_fs, t_fe)
        # Synchronous draw — see comment in _refresh_panel_styles.
        debug_draw("_update_panel_for_pair",
                    f"idx={idx} before draw()")
        self._canvas.draw()
        debug_draw("_update_panel_for_pair",
                    f"idx={idx} after draw()")

    def _open_panel_popout(self, idx: int) -> None:
        """Open the pair at `idx` in a Toplevel window with a single
        large subplot and the matplotlib NavigationToolbar (pan, zoom,
        save). Edge-drag-to-edit and right-click-to-revert work inside
        the popout. Closing the popout refreshes the grid panel with
        any changes made."""
        if idx >= len(self.session.pairs):
            return
        if self.session._raw_trace is None:
            return
        p = self.session.pairs[idx]
        if p._enzyme_seg is None or p._blank_seg is None:
            return

        win = tk.Toplevel(self.app.root)
        win.title(f"{self.session.name} — pair #{p.pair}")
        win.geometry("960x620")

        fig = Figure(figsize=(9, 5.5), dpi=110)
        ax = fig.add_subplot(111)

        t_full, a_full = self.session._raw_trace
        s_e, e_e = p._enzyme_seg
        s_b, e_b = p._blank_seg
        xs_e = t_full[s_e:e_e]
        ys_e = a_full[s_e:e_e]
        xs_b = t_full[s_b:e_b]
        ys_b_raw = a_full[s_b:e_b]
        # Shift blank for visual comparison
        offset = ys_e[0] - ys_b_raw[0] if len(ys_e) and len(ys_b_raw) else 0.0
        ys_b = ys_b_raw + offset
        enz_line, = ax.plot(xs_e, ys_e, color="tab:red",
                            linewidth=0.9, label="enzyme")
        blank_line, = ax.plot(xs_b, ys_b, color="tab:blue",
                              linewidth=0.9, alpha=0.7,
                              label="blank (shifted)")

        # Enzyme fit
        t_fs = p.t_enzyme_start
        t_fe = p.t_enzyme_start + p.window_enzyme_s / 60.0
        anchor_idx = int(np.searchsorted(xs_e, t_fs))
        anchor_idx = min(max(anchor_idx, 0), len(ys_e) - 1)
        a_anchor = ys_e[anchor_idx] if len(ys_e) else 0.0
        fit_line, = ax.plot(
            [t_fs, t_fe],
            [a_anchor, a_anchor + p.slope_enzyme * (t_fe - t_fs)],
            color="black", linewidth=2.0, label="enzyme fit")

        # Blank fit (dashed) for reference
        if p._blank_fit_seg is not None:
            sb_fit, eb_fit = p._blank_fit_seg
            xs_bf_lo = float(t_full[sb_fit])
            xs_bf_hi = float(t_full[eb_fit - 1])
            ai_bl = int(np.searchsorted(xs_b, xs_bf_lo))
            ai_bl = min(max(ai_bl, 0), len(ys_b) - 1)
            a_anc_bl = ys_b[ai_bl]
            ax.plot([xs_bf_lo, xs_bf_hi],
                    [a_anc_bl,
                     a_anc_bl + p.slope_blank * (xs_bf_hi - xs_bf_lo)],
                    color="dimgrey", linewidth=1.4, linestyle="--",
                    label="blank fit")

        # Shaded fit window
        shade = ax.axvspan(t_fs, t_fe, color="black", alpha=0.07,
                            zorder=-2)

        ax.set_xlabel("Time (min)")
        ax.set_ylabel("Absorbance")
        ax.set_title(self._panel_title_for(p))
        ax.legend(loc="best", fontsize=8)
        ax.grid(alpha=0.3)
        fig.tight_layout()

        from matplotlib.backends.backend_tkagg import (
            FigureCanvasTkAgg, NavigationToolbar2Tk)
        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        toolbar_frame = ttk.Frame(win)
        toolbar_frame.pack(fill="x")
        toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
        toolbar.update()
        canvas.draw()

        # Edge-drag editing inside the popout
        state = {"drag": None}
        zone_min = 0.3 / 60.0  # minutes; same hit-zone as in the grid

        def edge_at(x):
            if x is None:
                return None
            if abs(x - t_fs) <= zone_min:
                return "left"
            if abs(x - (p.t_enzyme_start + p.window_enzyme_s / 60.0)
                   ) <= zone_min:
                return "right"
            return None

        def on_press(event):
            if event.inaxes is not ax:
                return
            # Skip if user is in pan/zoom mode
            mode = getattr(toolbar, "mode", "")
            if mode and mode != "":
                return
            if event.button == 3:
                # Right-click: revert to auto
                if p.manually_adjusted:
                    revert_pair_to_auto(p)
                    refresh_popout()
                    recompute_session(self.session, self.app.state.params)
                    self.update_row_displays()
                    self._refresh_panel_styles()
                    self._refresh_panel_titles()
                    self.update_result_label()
                    self._update_panel_for_pair(idx)
                    self.app.recomputed_one(
                        self.session, refresh_self=False)
                return
            if event.button != 1 or event.xdata is None:
                return
            ed = edge_at(event.xdata)
            if ed is None:
                return
            state["drag"] = ed

        def on_motion(event):
            if state["drag"] is None or event.xdata is None:
                return
            new_x = float(event.xdata)
            cur_start = p.t_enzyme_start
            cur_end = p.t_enzyme_start + p.window_enzyme_s / 60.0
            if state["drag"] == "left":
                cur_start = max(float(xs_e[0]),
                                 min(new_x, cur_end - 1.0 / 60.0))
            else:
                cur_end = min(float(xs_e[-1]),
                              max(new_x, cur_start + 1.0 / 60.0))
            # Live preview: just move the shade and fit-line endpoints
            shade.set_xy([
                [cur_start, 0], [cur_start, 1],
                [cur_end, 1], [cur_end, 0],
                [cur_start, 0],
            ])
            ai = int(np.searchsorted(xs_e, cur_start))
            ai = min(max(ai, 0), len(ys_e) - 1)
            anch = ys_e[ai] if len(ys_e) else 0.0
            # tentative slope from a quick linreg on the window
            m_idx = (xs_e >= cur_start) & (xs_e <= cur_end)
            if m_idx.sum() >= 3:
                slope_tmp = float(np.polyfit(
                    xs_e[m_idx], ys_e[m_idx], 1)[0])
            else:
                slope_tmp = p.slope_enzyme
            fit_line.set_data(
                [cur_start, cur_end],
                [anch, anch + slope_tmp * (cur_end - cur_start)])
            canvas.draw()

        def on_release(event):
            if state["drag"] is None:
                return
            cur_start = p.t_enzyme_start
            cur_end = p.t_enzyme_start + p.window_enzyme_s / 60.0
            if event.xdata is not None:
                if state["drag"] == "left":
                    cur_start = max(float(xs_e[0]),
                                     min(float(event.xdata),
                                         cur_end - 1.0 / 60.0))
                else:
                    cur_end = min(float(xs_e[-1]),
                                  max(float(event.xdata),
                                      cur_start + 1.0 / 60.0))
            state["drag"] = None
            ok = refit_pair_window(
                p, t_full, a_full, cur_start, cur_end,
                r2_thresh=self.app.state.params.r2_thresh)
            if ok:
                refresh_popout()
                recompute_session(self.session, self.app.state.params)
                self.update_row_displays()
                self._refresh_panel_styles()
                self._refresh_panel_titles()
                self.update_result_label()
                self._update_panel_for_pair(idx)
                self.app.recomputed_one(
                    self.session, refresh_self=False)

        def refresh_popout():
            """Repaint the popout's fit line, shade, and title from p."""
            t_fs2 = p.t_enzyme_start
            t_fe2 = p.t_enzyme_start + p.window_enzyme_s / 60.0
            ai2 = int(np.searchsorted(xs_e, t_fs2))
            ai2 = min(max(ai2, 0), len(ys_e) - 1)
            anc2 = ys_e[ai2] if len(ys_e) else 0.0
            fit_line.set_data(
                [t_fs2, t_fe2],
                [anc2, anc2 + p.slope_enzyme * (t_fe2 - t_fs2)])
            shade.set_xy([
                [t_fs2, 0], [t_fs2, 1],
                [t_fe2, 1], [t_fe2, 0],
                [t_fs2, 0],
            ])
            ax.set_title(self._panel_title_for(p))
            canvas.draw()

        canvas.mpl_connect("button_press_event", on_press)
        canvas.mpl_connect("motion_notify_event", on_motion)
        canvas.mpl_connect("button_release_event", on_release)

        def on_close():
            win.destroy()

        ttk.Button(toolbar_frame, text="Close", command=on_close
                   ).pack(side="right", padx=6)
        win.protocol("WM_DELETE_WINDOW", on_close)

    # ------------------------------------------------------------------
    # Table
    # ------------------------------------------------------------------

    def _make_row(self, tbl, row_idx: int, p: PairRow):
        widgets = {}
        ttk.Label(tbl, text=str(p.pair)).grid(
            row=row_idx, column=0, padx=4, pady=1, sticky="w")
        ttk.Label(tbl, text=fmt(p.t_enzyme_start, 4)).grid(
            row=row_idx, column=1, padx=4, pady=1, sticky="w")
        ttk.Label(tbl, text=fmt(p.slope_blank, 4)).grid(
            row=row_idx, column=2, padx=4, pady=1, sticky="w")
        slope_enz_lbl = ttk.Label(tbl, text=fmt(p.slope_enzyme, 4))
        slope_enz_lbl.grid(row=row_idx, column=3, padx=4, pady=1, sticky="w")
        widgets["slope_enz_lbl"] = slope_enz_lbl
        slope_net_lbl = ttk.Label(tbl, text=fmt(p.slope_net, 4))
        slope_net_lbl.grid(row=row_idx, column=4, padx=4, pady=1, sticky="w")
        widgets["slope_net_lbl"] = slope_net_lbl
        r2_lbl = ttk.Label(tbl, text=fmt(p.r2_enzyme, 4))
        r2_lbl.grid(row=row_idx, column=5, padx=4, pady=1, sticky="w")
        widgets["r2_lbl"] = r2_lbl
        meets_lbl = ttk.Label(tbl, text=("yes" if p.meets_r2 else "NO"),
                              foreground=("black" if p.meets_r2 else "red"))
        meets_lbl.grid(row=row_idx, column=6, padx=4, pady=1, sticky="w")
        widgets["meets_lbl"] = meets_lbl
        win_lbl = ttk.Label(tbl, text=fmt(p.window_enzyme_s, 3))
        win_lbl.grid(row=row_idx, column=7, padx=4, pady=1, sticky="w")
        widgets["win_lbl"] = win_lbl
        manual_lbl = ttk.Label(tbl, text=("✎" if p.manually_adjusted else ""),
                               foreground="#1f77b4")
        manual_lbl.grid(row=row_idx, column=8, padx=4, pady=1, sticky="w")
        widgets["manual_lbl"] = manual_lbl

        c_var = tk.StringVar(value=fmt(p.concentration_uM, 6))
        c_entry = ttk.Entry(tbl, width=8, textvariable=c_var)
        c_entry.grid(row=row_idx, column=9, padx=4, pady=1, sticky="w")
        c_entry.bind("<FocusOut>",
                     lambda _e, pp=p, var=c_var: self._on_conc(pp, var))
        c_entry.bind("<Return>",
                     lambda _e, pp=p, var=c_var: self._on_conc(pp, var))
        widgets["c_var"] = c_var

        v_lbl = ttk.Label(tbl, text="")
        v_lbl.grid(row=row_idx, column=10, padx=4, pady=1, sticky="w")
        widgets["v_lbl"] = v_lbl
        k_lbl = ttk.Label(tbl, text="")
        k_lbl.grid(row=row_idx, column=11, padx=4, pady=1, sticky="w")
        widgets["k_lbl"] = k_lbl

        inc_var = tk.BooleanVar(value=p.include)
        inc_cb = ttk.Checkbutton(
            tbl, variable=inc_var,
            command=lambda pp=p, var=inc_var: self._on_include(pp, var))
        inc_cb.grid(row=row_idx, column=12, padx=4, pady=1, sticky="w")
        widgets["inc_var"] = inc_var

        self._row_widgets.append(widgets)

    def _on_conc(self, p: PairRow, var: tk.StringVar):
        v = parse_float(var.get())
        if v <= 0 or not np.isfinite(v):
            p.concentration_uM = float("nan")
        else:
            p.concentration_uM = v
        var.set(fmt(p.concentration_uM, 6))
        recompute_session(self.session, self.app.state.params)
        self.update_row_displays()
        self._refresh_panel_titles()
        self.update_result_label()
        self.app.recomputed_one(self.session, refresh_self=False)

    def _on_include(self, p: PairRow, var: tk.BooleanVar):
        p.include = bool(var.get())
        recompute_session(self.session, self.app.state.params)
        self.update_row_displays()
        self._refresh_panel_styles()
        self._refresh_panel_titles()
        self.update_result_label()
        self.app.recomputed_one(self.session, refresh_self=False)

    def _set_all_included(self, value: bool):
        for p, w in zip(self.session.pairs, self._row_widgets):
            p.include = value
            w["inc_var"].set(value)
        recompute_session(self.session, self.app.state.params)
        self.update_row_displays()
        self._refresh_panel_styles()
        self._refresh_panel_titles()
        self.update_result_label()
        self.app.recomputed_one(self.session, refresh_self=False)

    def _reset_concentrations(self):
        if len(self.session.pairs) == len(DEFAULT_CONCENTRATIONS_UM):
            for p, c, w in zip(self.session.pairs,
                               DEFAULT_CONCENTRATIONS_UM,
                               self._row_widgets):
                p.concentration_uM = c
                w["c_var"].set(fmt(c, 6))
        else:
            messagebox.showinfo(
                "Cannot reset",
                f"This file has {len(self.session.pairs)} pairs but the "
                f"default concentration list has "
                f"{len(DEFAULT_CONCENTRATIONS_UM)} entries.")
            return
        recompute_session(self.session, self.app.state.params)
        self.update_row_displays()
        self._refresh_panel_titles()
        self.update_result_label()
        self.app.recomputed_one(self.session, refresh_self=False)

    def update_row_displays(self):
        """Recompute and refresh all per-row labels: slope/R²/window
        labels (which change when a window is manually adjusted) plus
        the v / kobs labels (which change when [E], ε, or [S] change)."""
        for p, w in zip(self.session.pairs, self._row_widgets):
            w["slope_enz_lbl"].configure(text=fmt(p.slope_enzyme, 4))
            w["slope_net_lbl"].configure(text=fmt(p.slope_net, 4))
            w["r2_lbl"].configure(text=fmt(p.r2_enzyme, 4))
            w["meets_lbl"].configure(
                text=("yes" if p.meets_r2 else "NO"),
                foreground=("black" if p.meets_r2 else "red"))
            w["win_lbl"].configure(text=fmt(p.window_enzyme_s, 3))
            w["manual_lbl"].configure(
                text=("✎" if p.manually_adjusted else ""))
            if p.include and np.isfinite(p.concentration_uM) \
                    and p.concentration_uM > 0:
                v = velocity_uM_per_s(p.slope_net, self.session.epsilon_M1cm1)
                k = kobs_per_s(v, self.session.enzyme_conc_uM)
                w["v_lbl"].configure(text=fmt(v, 4))
                w["k_lbl"].configure(text=fmt(k, 4))
            else:
                w["v_lbl"].configure(text="—")
                w["k_lbl"].configure(text="—")

    def update_result_label(self):
        s = self.session
        if not np.isfinite(s.kcat_over_km):
            txt = "kcat/KM:  (need at least 2 valid points)"
        else:
            method = s.kcat_method_used
            if method == "bootstrap" and np.isfinite(s.kcat_ci_lo) \
                    and np.isfinite(s.kcat_ci_hi):
                kcat_txt = (f"kcat/KM = {s.kcat_over_km:.4g} M⁻¹s⁻¹  "
                            f"[95% CI {s.kcat_ci_lo:.4g}, "
                            f"{s.kcat_ci_hi:.4g}]  (bootstrap)")
            elif method == "per_replicate" and np.isfinite(s.kcat_over_km_se):
                kcat_txt = (f"kcat/KM = {s.kcat_over_km:.4g} "
                            f"± {s.kcat_over_km_se:.2g} M⁻¹s⁻¹  "
                            f"(SD across replicates)")
            elif method == "weighted" and np.isfinite(s.kcat_over_km_se):
                kcat_txt = (f"kcat/KM = {s.kcat_over_km:.4g} "
                            f"± {s.kcat_over_km_se:.2g} M⁻¹s⁻¹  "
                            f"(weighted SE)")
            else:
                kcat_txt = f"kcat/KM = {s.kcat_over_km:.4g} M⁻¹s⁻¹"
            txt = (f"{kcat_txt}    "
                   f"R² = {s.fit_r2:.4f}    "
                   f"intercept = {s.fit_intercept:.4g} s⁻¹    "
                   f"n = {s.n_used}")
        self._result_lbl.configure(text=txt)

    def refresh_external(self):
        """Called by the App when ε or [E] change from the Setup tab."""
        if not self._lazy_built:
            return  # nothing to refresh visually; will rebuild on first click
        self.update_row_displays()
        self._refresh_panel_titles()
        self.update_result_label()


class PlotTab(ttk.Frame):
    """Overlay kobs vs [S] for all enzymes, with linear fits."""

    def __init__(self, master, app: "App"):
        super().__init__(master)
        self.app = app
        self.fig = Figure(figsize=(8, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        toolbar = NavigationToolbar2Tk(self.canvas, self)
        toolbar.update()
        ttk.Button(self, text="Refresh plot",
                   command=self.redraw).pack(pady=4)

    def redraw(self):
        self.ax.clear()
        cmap = matplotlib.colormaps.get_cmap("tab10")
        any_drawn = False
        for i, s in enumerate(self.app.state.sessions):
            xs, ys = [], []
            for p in s.pairs:
                if not p.include:
                    continue
                if not (np.isfinite(p.concentration_uM)
                        and p.concentration_uM > 0):
                    continue
                v = velocity_uM_per_s(p.slope_net, s.epsilon_M1cm1)
                k = kobs_per_s(v, s.enzyme_conc_uM)
                if not np.isfinite(k):
                    continue
                xs.append(p.concentration_uM)
                ys.append(k)
            if not xs:
                continue
            color = cmap(i % 10)
            label = (f"{s.name}  (kcat/KM = {fmt(s.kcat_over_km, 3)} "
                     f"M⁻¹s⁻¹)")
            self.ax.scatter(xs, ys, color=color, s=30, label=label,
                            edgecolors="black", linewidths=0.5)
            if np.isfinite(s.kcat_over_km):
                xx = np.linspace(0, max(xs) * 1.05, 50)
                yy = (s.kcat_over_km * 1e-6) * xx + s.fit_intercept
                self.ax.plot(xx, yy, color=color, linewidth=1.2, alpha=0.8)
            any_drawn = True
        self.ax.set_xlabel("[S] (µM)")
        self.ax.set_ylabel("kobs (s⁻¹)")
        self.ax.set_title("kobs vs [S]   (slope × 10⁶ = kcat/KM in M⁻¹s⁻¹)")
        if any_drawn:
            self.ax.legend(fontsize=7, loc="best")
        self.ax.grid(alpha=0.3)
        self.fig.tight_layout()
        self.canvas.draw_idle()


# ---------------------------------------------------------------------------
# Main app
# ---------------------------------------------------------------------------

class App:
    def __init__(self, root: tk.Tk, initial_folder: Optional[Path] = None):
        self.root = root
        self.state = AppState()
        # If the user has saved their preferred parameter values to
        # kinetics_defaults.json (next to this script), apply them on
        # top of the hardcoded defaults. Missing or malformed defaults
        # are ignored — the factory values stay in place.
        self.state.params = make_params_with_user_defaults()

        root.title("Enzyme kinetics — batch analyser")
        root.geometry("1200x780")

        # Top toolbar
        top = ttk.Frame(root)
        top.pack(fill="x", padx=8, pady=6)
        ttk.Button(top, text="Open folder…", command=self.open_folder
                   ).pack(side="left")
        self.folder_lbl = ttk.Label(top, text="(no folder)")
        self.folder_lbl.pack(side="left", padx=10)
        ttk.Button(top, text="Re-extract slopes", command=self.reload_folder
                   ).pack(side="left", padx=4)
        ttk.Button(top, text="Export results…", command=self.export
                   ).pack(side="right", padx=2)

        # Notebook
        self.nb = ttk.Notebook(root)
        self.nb.pack(fill="both", expand=True, padx=8, pady=4)
        # Lazy build: when the user clicks a file tab for the first
        # time, construct its heavy contents (raw trace, figures,
        # region-selection plot). This keeps folder-load fast on many
        # files.
        self.nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self.setup_tab = SetupTab(self.nb, self)
        self.nb.add(self.setup_tab, text="Setup")
        self.params_tab = ParametersTab(self.nb, self)
        self.nb.add(self.params_tab, text="Parameters")
        self.plot_tab = PlotTab(self.nb, self)
        self.nb.add(self.plot_tab, text="Plot")

        self.file_tabs: list[FileTab] = []

        # Status bar
        self.status = ttk.Label(root, text="Ready.", anchor="w",
                                relief="sunken")
        self.status.pack(fill="x", side="bottom")

        if initial_folder is not None and initial_folder.is_dir():
            self.root.after(50, lambda: self.load_folder(initial_folder))

    # -- folder I/O ------------------------------------------------------

    def open_folder(self):
        d = filedialog.askdirectory(title="Choose folder with Cary CSV files")
        if d:
            self.load_folder(Path(d))

    def _on_tab_changed(self, event=None):
        """Fire when the user clicks a tab. Lazy-builds FileTab content
        the first time it's selected."""
        try:
            current = self.nb.select()
        except Exception:
            return
        for t in self.file_tabs:
            if str(t) == current and not getattr(t, "_lazy_built", True):
                self.status.configure(text=f"Loading {t.session.name}…")
                self.root.update_idletasks()
                t.lazy_build()
                self.status.configure(text="Ready.")
                break

    def reload_folder(self):
        if self.state.folder is not None:
            self.load_folder(self.state.folder)

    def load_folder(self, folder: Path):
        self.state.folder = folder
        self.folder_lbl.configure(text=str(folder))

        # Clear existing file tabs
        for t in self.file_tabs:
            self.nb.forget(t)
        self.file_tabs.clear()
        self.state.sessions.clear()

        files = discover_csv_files(folder)
        if not files:
            self.status.configure(text=f"No CSV files in {folder}.")
            self.setup_tab.rebuild()
            self.plot_tab.redraw()
            return

        self.status.configure(text=f"Loading {len(files)} files…")
        self.root.update_idletasks()

        # Try to read the sidecar for cached confirmations
        sidecar = load_sidecar(folder)
        sidecar_files = sidecar.get("files", {}) if sidecar else {}

        n_restored = 0
        for k, p in enumerate(files, start=1):
            self.status.configure(
                text=f"Loading [{k}/{len(files)}]: {p.name}")
            self.root.update_idletasks()
            sess = FileSession(
                path=p, name=p.stem,
                enzyme_conc_uM=DEFAULT_ENZYME_CONC_UM,
                epsilon_M1cm1=DEFAULT_EPSILON_M1CM1,
            )
            cached = sidecar_files.get(p.name)
            restored_from_cache = False
            if cached and cached.get("pairs_confirmed"):
                try:
                    sess.segments = [SegmentSpec.from_dict(d)
                                     for d in cached.get("segments", [])]
                    sess.enzyme_conc_uM = float(cached.get(
                        "enzyme_conc_uM", DEFAULT_ENZYME_CONC_UM))
                    sess.epsilon_M1cm1 = float(cached.get(
                        "epsilon_M1cm1", DEFAULT_EPSILON_M1CM1))

                    snapshots = cached.get("pair_snapshots")
                    if snapshots:
                        # v2 path: restore computed slopes verbatim, no
                        # recomputation. The numbers you see are exactly
                        # the numbers that were saved.
                        sess.pairs = restore_pairs_from_snapshots(snapshots)
                        sess.error = None
                    else:
                        # v1 fallback: recompute from segments, apply
                        # per-pair overrides on top.
                        pairs, err = build_pairs_from_segments(
                            p, sess.segments, params=self.state.params)
                        sess.pairs = pairs
                        sess.error = err
                        overrides = cached.get("pair_overrides", [])
                        for r, ov in zip(sess.pairs, overrides):
                            if ov.get("concentration_uM") is not None:
                                r.concentration_uM = float(
                                    ov["concentration_uM"])
                            r.include = bool(ov.get("include", True))
                        if any(ov.get("manual_window") for ov in overrides):
                            try:
                                df = kp.load_cary_csv(p)
                                sess._raw_trace = (df["time_min"].to_numpy(),
                                                    df["abs"].to_numpy())
                                t_full, a_full = sess._raw_trace
                                for r, ov in zip(sess.pairs, overrides):
                                    mw = ov.get("manual_window")
                                    if mw and len(mw) == 2:
                                        refit_pair_window(
                                            r, t_full, a_full,
                                            float(mw[0]), float(mw[1]),
                                            r2_thresh=self.state.params.r2_thresh)
                            except Exception:
                                pass
                    sess.pairs_confirmed = True
                    restored_from_cache = True
                    n_restored += 1
                except Exception as exc:
                    sess.error = f"Sidecar restore failed: {exc}"
                    sess.segments = []
                    sess.pairs_confirmed = False
            if not restored_from_cache:
                # Auto-detect segments as a starting suggestion. Pairs
                # are NOT built until the user confirms in step 1.
                segs, err = auto_detect_segments(
                    p, params=self.state.params)
                sess.segments = segs
                sess.error = err
                sess.pairs_confirmed = False
                # We do build a tentative pair list so step 2 isn't
                # empty if the user clicks straight into it; the
                # _switch_view guard will redirect them anyway.
                if segs:
                    pairs, _ = build_pairs_from_segments(
                        p, segs, params=self.state.params)
                    sess.pairs = pairs

            recompute_session(sess, self.state.params)
            self.state.sessions.append(sess)

        # Rebuild Setup, then add a tab per file (inserted before Plot)
        self.setup_tab.rebuild()
        plot_idx = self.nb.index(self.plot_tab)
        for sess in self.state.sessions:
            tab = FileTab(self.nb, self, sess)
            self.nb.insert(plot_idx, tab, text=sess.name)
            plot_idx += 1
            self.file_tabs.append(tab)
        self.plot_tab.redraw()
        msg = f"Loaded {len(files)} files."
        if n_restored:
            msg += f" Restored {n_restored} confirmed variant(s) from sidecar."
        msg += " Review regions in step 1, then advance to step 2."
        self.status.configure(text=msg)

    # -- sidecar persistence ---------------------------------------------

    def on_session_confirmed(self, sess: FileSession):
        """Called by FileTab when the user confirms regions in step 1.
        Persists the sidecar."""
        if self.state.folder is not None:
            save_sidecar(self.state.folder, self.state.sessions)
        # Refresh aggregate views
        self.setup_tab.refresh_results()
        self.plot_tab.redraw()

    # -- recompute hooks -------------------------------------------------

    def reextract_all(self):
        """Re-run slope extraction with current PipelineParams.

        For each session:
        - If `pairs_confirmed=True`, the segments are kept verbatim and
          only the per-pair fits are recomputed with the new params
          (preserving the user's manual region edits).
        - If `pairs_confirmed=False`, segmentation is re-run from scratch
          and a fresh segment list replaces the old one.

        Per-pair manual fit-window edits are PRESERVED across re-extract
        — they're re-applied on top of the new auto-fits. To clear a
        manual edit, right-click the panel in step 2.

        Per-pair [S], include flags, and per-variant [E]/ε are also
        preserved when the pair count matches.
        """
        if not self.state.sessions:
            return
        snap: dict[Path, dict] = {}
        for s in self.state.sessions:
            snap[s.path] = dict(
                enzyme_conc_uM=s.enzyme_conc_uM,
                epsilon_M1cm1=s.epsilon_M1cm1,
                pairs=[
                    {
                        "concentration_uM": p.concentration_uM,
                        "include": p.include,
                        "manual_window": (
                            (p.t_enzyme_start,
                             p.t_enzyme_start + p.window_enzyme_s / 60.0)
                            if p.manually_adjusted else None
                        ),
                    }
                    for p in s.pairs
                ],
                confirmed=s.pairs_confirmed,
                segments=list(s.segments),
            )

        # Remember which file tab is currently active so we can reselect
        # it after the rebuild.
        try:
            active_idx = self.nb.index(self.nb.select())
        except Exception:
            active_idx = None
        active_session_path: Optional[Path] = None
        if active_idx is not None:
            for tab, sess in zip(self.file_tabs, self.state.sessions):
                if self.nb.index(tab) == active_idx:
                    active_session_path = sess.path
                    break

        self.status.configure(text="Re-extracting slopes…")
        self.root.update_idletasks()

        n_manual_preserved = 0
        for s in self.state.sessions:
            if snap[s.path]["confirmed"]:
                rows, err = build_pairs_from_segments(
                    s.path, s.segments, params=self.state.params)
            else:
                segs, err1 = auto_detect_segments(
                    s.path, params=self.state.params)
                s.segments = segs
                rows, err = build_pairs_from_segments(
                    s.path, segs, params=self.state.params)
                err = err1 or err
            s.pairs = rows
            s.error = err
            s.enzyme_conc_uM = snap[s.path]["enzyme_conc_uM"]
            s.epsilon_M1cm1 = snap[s.path]["epsilon_M1cm1"]
            old = snap[s.path]["pairs"]
            if len(old) == len(s.pairs):
                # Reapply per-pair concentration and include
                for r, info in zip(s.pairs, old):
                    c = info["concentration_uM"]
                    if np.isfinite(c) and c > 0:
                        r.concentration_uM = c
                    r.include = info["include"]
                # Reapply manual fit windows (after the new auto fit
                # has been computed; manual takes precedence)
                if s._raw_trace is None:
                    try:
                        df = kp.load_cary_csv(s.path)
                        s._raw_trace = (df["time_min"].to_numpy(),
                                        df["abs"].to_numpy())
                    except Exception:
                        pass
                if s._raw_trace is not None:
                    t_full, a_full = s._raw_trace
                    for r, info in zip(s.pairs, old):
                        mw = info["manual_window"]
                        if mw is not None:
                            ok = refit_pair_window(
                                r, t_full, a_full,
                                float(mw[0]), float(mw[1]),
                                r2_thresh=self.state.params.r2_thresh)
                            if ok:
                                n_manual_preserved += 1
            recompute_session(s, self.state.params)

        # Rebuild file tabs (figures need to be rebuilt)
        for t in self.file_tabs:
            self.nb.forget(t)
        self.file_tabs.clear()
        plot_idx = self.nb.index(self.plot_tab)
        new_active_tab = None
        for sess in self.state.sessions:
            tab = FileTab(self.nb, self, sess)
            self.nb.insert(plot_idx, tab, text=sess.name)
            plot_idx += 1
            self.file_tabs.append(tab)
            if active_session_path is not None and sess.path == active_session_path:
                new_active_tab = tab
        self.setup_tab.rebuild()
        self.plot_tab.redraw()
        if self.state.folder is not None:
            save_sidecar(self.state.folder, self.state.sessions)

        # Reselect the FileTab the user was viewing so they immediately
        # see the new fits, rather than landing on the Setup tab.
        if new_active_tab is not None:
            self.nb.select(new_active_tab)

        msg = "Re-extraction complete."
        if n_manual_preserved:
            msg += f" {n_manual_preserved} manual fit window(s) preserved."
        self.status.configure(text=msg)

    def recomputed_one(self, sess: FileSession, refresh_self: bool = True):
        """Called when a single session has changed.

        `refresh_self=False` means the caller has already updated its own
        tab in place and only the Setup + Plot tabs need refreshing.
        """
        if refresh_self:
            for t in self.file_tabs:
                if t.session is sess:
                    t.refresh_external()
                    break
        self.setup_tab.refresh_results()
        self.plot_tab.redraw()
        if self.state.folder is not None:
            save_sidecar(self.state.folder, self.state.sessions)

    def recomputed_all(self):
        for t in self.file_tabs:
            t.refresh_external()
        self.setup_tab.refresh_results()
        self.plot_tab.redraw()
        if self.state.folder is not None:
            save_sidecar(self.state.folder, self.state.sessions)

    # -- export ----------------------------------------------------------

    def export(self):
        if not self.state.sessions:
            messagebox.showinfo("Nothing to export", "Load a folder first.")
            return
        d = filedialog.askdirectory(title="Choose output folder")
        if not d:
            return
        out = Path(d)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Missing dependency",
                "Excel export needs the openpyxl package. Install it with:\n\n"
                "    conda install openpyxl\n  or\n    pip install openpyxl")
            return

        xlsx_path = out / "kinetics_results.xlsx"
        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2A5D8F")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        excluded_fill = PatternFill("solid", fgColor="F2F2F2")

        def write_header(ws, row, cols):
            for ci, name in enumerate(cols, 1):
                c = ws.cell(row=row, column=ci, value=name)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")
                c.border = border

        def autosize(ws):
            for col in ws.columns:
                width = max((len(str(c.value)) if c.value is not None else 0
                             for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = \
                    min(max(width + 2, 10), 28)

        # ---------------- Summary sheet ----------------
        ws = wb.active
        ws.title = "Summary"
        cols = ["variant", "n_pairs", "n_used", "method",
                "kcat_over_km (M^-1 s^-1)",
                "uncertainty (M^-1 s^-1)",
                "uncertainty_label",
                "CI95_lo (M^-1 s^-1)",
                "CI95_hi (M^-1 s^-1)",
                "intercept (s^-1)", "fit_R2",
                "[E] (uM)", "epsilon (M^-1 cm^-1)"]
        write_header(ws, 1, cols)
        for ri, s in enumerate(self.state.sessions, 2):
            method = getattr(s, "kcat_method_used", "weighted")
            if method == "per_replicate":
                unc_label = "SD across replicates"
            elif method == "weighted":
                unc_label = "weighted SE"
            elif method == "bootstrap":
                unc_label = "see CI95 columns"
            else:
                unc_label = ""
            ws.cell(row=ri, column=1, value=s.name)
            ws.cell(row=ri, column=2, value=len(s.pairs))
            ws.cell(row=ri, column=3, value=s.n_used)
            ws.cell(row=ri, column=4, value=method)
            ws.cell(row=ri, column=5,
                    value=float(s.kcat_over_km)
                    if np.isfinite(s.kcat_over_km) else None)
            ws.cell(row=ri, column=6,
                    value=float(s.kcat_over_km_se)
                    if np.isfinite(s.kcat_over_km_se) else None)
            ws.cell(row=ri, column=7, value=unc_label)
            ws.cell(row=ri, column=8,
                    value=float(s.kcat_ci_lo)
                    if np.isfinite(s.kcat_ci_lo) else None)
            ws.cell(row=ri, column=9,
                    value=float(s.kcat_ci_hi)
                    if np.isfinite(s.kcat_ci_hi) else None)
            ws.cell(row=ri, column=10,
                    value=float(s.fit_intercept)
                    if np.isfinite(s.fit_intercept) else None)
            ws.cell(row=ri, column=11,
                    value=float(s.fit_r2)
                    if np.isfinite(s.fit_r2) else None)
            ws.cell(row=ri, column=12, value=float(s.enzyme_conc_uM))
            ws.cell(row=ri, column=13, value=float(s.epsilon_M1cm1))
            # Number formats
            for col in (5, 6, 8, 9):
                ws.cell(row=ri, column=col).number_format = "0.0"
            ws.cell(row=ri, column=10).number_format = "0.000E+00"
            ws.cell(row=ri, column=11).number_format = "0.0000"
            ws.cell(row=ri, column=12).number_format = "0.0000"
            ws.cell(row=ri, column=13).number_format = "0.0"
        autosize(ws)
        ws.freeze_panes = "A2"

        # ---------------- Setup sheet ----------------
        ws = wb.create_sheet("Setup")
        write_header(ws, 1, ["variant", "[E] (uM)", "epsilon (M^-1 cm^-1)"])
        for ri, s in enumerate(self.state.sessions, 2):
            ws.cell(row=ri, column=1, value=s.name)
            ws.cell(row=ri, column=2, value=float(s.enzyme_conc_uM))
            ws.cell(row=ri, column=3, value=float(s.epsilon_M1cm1))
            ws.cell(row=ri, column=2).number_format = "0.0000"
            ws.cell(row=ri, column=3).number_format = "0.0"
        autosize(ws)
        ws.freeze_panes = "A2"

        # ---------------- One sheet per variant ----------------
        cols = ["pair", "include", "concentration (uM)",
                "t_start (min)", "window (s)",
                "slope_blank (AU/min)", "slope_enzyme (AU/min)",
                "slope_net (AU/min)", "R2_enzyme", "meets_R2",
                "v (uM/s)", "kobs (s^-1)",
                "manually_adjusted"]
        for s in self.state.sessions:
            # Excel sheet names: max 31 chars, no : \ / ? * [ ]
            safe = "".join("_" if c in ':\\/?*[]' else c
                           for c in s.name)[:31]
            ws = wb.create_sheet(safe)
            write_header(ws, 1, cols)
            for ri, p in enumerate(s.pairs, 2):
                v = (velocity_uM_per_s(p.slope_net, s.epsilon_M1cm1)
                     if np.isfinite(p.concentration_uM) else float("nan"))
                k = kobs_per_s(v, s.enzyme_conc_uM)
                row_vals = [
                    p.pair, bool(p.include),
                    float(p.concentration_uM)
                    if np.isfinite(p.concentration_uM) else None,
                    float(p.t_enzyme_start),
                    float(p.window_enzyme_s),
                    float(p.slope_blank),
                    float(p.slope_enzyme),
                    float(p.slope_net),
                    float(p.r2_enzyme)
                    if np.isfinite(p.r2_enzyme) else None,
                    bool(p.meets_r2),
                    float(v) if np.isfinite(v) else None,
                    float(k) if np.isfinite(k) else None,
                    bool(p.manually_adjusted),
                ]
                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    if not p.include:
                        c.fill = excluded_fill
                # Number formats
                ws.cell(row=ri, column=3).number_format = "0.00"
                ws.cell(row=ri, column=4).number_format = "0.000"
                ws.cell(row=ri, column=5).number_format = "0.0"
                for col in (6, 7, 8):
                    ws.cell(row=ri, column=col).number_format = "0.000000"
                ws.cell(row=ri, column=9).number_format = "0.0000"
                ws.cell(row=ri, column=11).number_format = "0.000000"
                ws.cell(row=ri, column=12).number_format = "0.000000"
            # Footer with variant summary
            footer_row = len(s.pairs) + 3
            method = getattr(s, "kcat_method_used", "weighted")
            ws.cell(row=footer_row, column=1,
                    value="method:").font = Font(bold=True)
            ws.cell(row=footer_row, column=2, value=method)
            ws.cell(row=footer_row + 1, column=1,
                    value="kcat/KM (M^-1 s^-1):").font = Font(bold=True)
            ws.cell(row=footer_row + 1, column=2,
                    value=float(s.kcat_over_km)
                    if np.isfinite(s.kcat_over_km) else None
                    ).number_format = "0.0"
            if method == "bootstrap":
                ws.cell(row=footer_row + 2, column=1,
                        value="95% CI lo:").font = Font(bold=True)
                ws.cell(row=footer_row + 2, column=2,
                        value=float(s.kcat_ci_lo)
                        if np.isfinite(s.kcat_ci_lo) else None
                        ).number_format = "0.0"
                ws.cell(row=footer_row + 3, column=1,
                        value="95% CI hi:").font = Font(bold=True)
                ws.cell(row=footer_row + 3, column=2,
                        value=float(s.kcat_ci_hi)
                        if np.isfinite(s.kcat_ci_hi) else None
                        ).number_format = "0.0"
                ws.cell(row=footer_row + 4, column=1,
                        value="n_used:").font = Font(bold=True)
                ws.cell(row=footer_row + 4, column=2, value=s.n_used)
            else:
                if method == "per_replicate":
                    se_label = "SD across replicates (M^-1 s^-1):"
                else:
                    se_label = "weighted SE (M^-1 s^-1):"
                ws.cell(row=footer_row + 2, column=1,
                        value=se_label).font = Font(bold=True)
                se_val = getattr(s, "kcat_over_km_se", float("nan"))
                ws.cell(row=footer_row + 2, column=2,
                        value=float(se_val) if np.isfinite(se_val) else None
                        ).number_format = "0.0"
                ws.cell(row=footer_row + 3, column=1,
                        value="n_used:").font = Font(bold=True)
                ws.cell(row=footer_row + 3, column=2, value=s.n_used)
            autosize(ws)
            ws.freeze_panes = "A2"

        try:
            wb.save(xlsx_path)
        except PermissionError:
            messagebox.showerror(
                "Cannot save",
                f"Could not write {xlsx_path}. The file may be open in "
                f"Excel — close it and try again.")
            return
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))
            return

        messagebox.showinfo(
            "Export complete",
            f"Wrote {xlsx_path.name} to {out}")
        self.status.configure(text=f"Exported to {xlsx_path}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

# Global debug logger (set up by main() when --debug-draw is on).
DEBUG_DRAW_LOG: Optional[object] = None


def debug_draw(site: str, detail: str = "") -> None:
    """If --debug-draw was passed, append a timestamped record of this
    redraw event to the log. No-op otherwise. Used to diagnose cases
    where manual edits don't update the canvas until the user clicks
    elsewhere — the log shows whether draw() actually fires immediately
    or is being deferred upstream."""
    if DEBUG_DRAW_LOG is None:
        return
    import time
    try:
        DEBUG_DRAW_LOG.write(
            f"{time.time():.4f}  {site:32s}  {detail}\n")
        DEBUG_DRAW_LOG.flush()
    except Exception:
        pass


def main(argv: list[str] | None = None) -> int:
    global DEBUG_DRAW_LOG
    args = sys.argv[1:] if argv is None else argv
    initial = None
    debug_draw_on = False
    for a in list(args):
        if a == "--debug-draw":
            debug_draw_on = True
            args.remove(a)
    if args:
        p = Path(args[0]).expanduser()
        if p.is_dir():
            initial = p

    if debug_draw_on:
        log_path = Path.cwd() / "kinetics_draw_debug.log"
        DEBUG_DRAW_LOG = log_path.open("w", encoding="utf-8")
        DEBUG_DRAW_LOG.write(
            f"# Kinetics GUI draw-debug log\n"
            f"# Each row: epoch_seconds  call_site  detail\n")
        DEBUG_DRAW_LOG.flush()
        print(f"Debug-draw logging to {log_path}")

    root = tk.Tk()
    try:
        App(root, initial_folder=initial)
        root.mainloop()
    except Exception:
        traceback.print_exc()
        return 1
    finally:
        if DEBUG_DRAW_LOG is not None:
            try:
                DEBUG_DRAW_LOG.close()
            except Exception:
                pass
    return 0


if __name__ == "__main__":
    sys.exit(main())